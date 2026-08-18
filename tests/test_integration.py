# -*- coding: utf-8 -*-
"""Testy integracyjne — współpraca parse / cache / excel / fetch (bez sieci)."""
from __future__ import annotations

from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

import scrape_footystats as sf


class TestParseIntegration:
    def test_parse_sample_html_yields_three_matches(self, sample_html):
        match_date, rows = sf.parse_matches(sample_html)
        assert match_date == "11/08"
        assert len(rows) == 3

    def test_parse_fields_and_absolute_link(self, sample_html):
        _, rows = sf.parse_matches(sample_html)
        first = rows[0]
        assert first["Kraj"] == "Европа"
        assert first["Liga"] == "Лига чемпионов УЕФА"
        assert first["Godzina"] == "18:00"
        assert first["Gospodarz"] == "Жальгирис"
        assert first["Gość"] == "Динамо Загреб"
        assert first["Kurs 1"] == "6.35"
        assert first["Status"] == "incomplete"
        assert first["Link"].startswith("https://footystats.org/")

    def test_country_stripped_from_league(self, sample_html):
        _, rows = sf.parse_matches(sample_html)
        england = [r for r in rows if r["Kraj"] == "Англия"]
        assert len(england) == 1
        assert england[0]["Liga"] == "Кубок Англии"


class TestCacheExcelIntegration:
    def test_cache_roundtrip_with_excel(self, tmp_paths, sample_html):
        match_date, rows = sf.parse_matches(sample_html)
        cache_path = sf.save_matches_cache(match_date, rows, "fixture")
        sf.save_html_cache(sample_html, "fixture")

        out = tmp_paths["root"] / "mecze.xlsx"
        sf.save_excel(rows, out)

        loaded = sf.load_matches_cache(ttl_minutes=30)
        assert loaded is not None
        assert loaded[0] == match_date
        assert len(loaded[1]) == 3
        assert cache_path.exists()
        assert tmp_paths["html_cache"].exists()

        wb = load_workbook(out)
        assert wb.active.max_row == 4
        assert wb.active.cell(2, 4).value == "18:00"

    def test_cookies_cache_roundtrip(self, tmp_paths):
        cookies = [
            {
                "name": "cf_clearance",
                "value": "abc",
                "domain": "footystats.org",
                "path": "/",
            }
        ]
        sf.save_cookies_cache(cookies)
        assert sf.load_cookies_cache(ttl_minutes=60) == cookies


class TestFetchIntegration:
    def test_fetch_html_uses_first_working_source(self, tmp_paths, sample_html):
        with (
            patch.object(sf, "fetch_with_playwright", return_value=sample_html),
            patch.object(sf, "fetch_with_curl_cffi", return_value=None) as curl_mock,
            patch.object(sf, "fetch_with_requests", return_value=None),
        ):
            html, source = sf.fetch_html("https://example.com")
        assert html == sample_html
        assert source == "playwright"
        curl_mock.assert_not_called()
        assert tmp_paths["html_cache"].exists()

    def test_fetch_html_fallback_chain(self, tmp_paths, sample_html):
        with (
            patch.object(sf, "fetch_with_playwright", return_value=None),
            patch.object(sf, "fetch_with_curl_cffi", return_value=sample_html),
            patch.object(sf, "fetch_with_requests", return_value=None),
        ):
            html, source = sf.fetch_html("https://example.com")
        assert source == "curl_cffi"
        assert "league" in html


class TestMainPipelineIntegration:
    def test_main_refresh_writes_excel_and_cache(self, tmp_paths, sample_html):
        sf.logger.handlers.clear()
        with (
            patch.object(sf, "fetch_with_playwright", return_value=sample_html),
            patch.object(sf, "fetch_with_curl_cffi", return_value=None),
            patch.object(sf, "fetch_with_requests", return_value=None),
            patch("sys.argv", ["scrape_footystats.py", "--refresh", "--log-level", "INFO"]),
        ):
            sf.main()

        assert tmp_paths["matches_cache"].exists()
        assert list(tmp_paths["root"].glob("footystats_mecze_*.xlsx"))
        sf.logger.handlers.clear()

    def test_main_second_run_uses_cache(self, tmp_paths, sample_html):
        sf.logger.handlers.clear()
        with (
            patch.object(sf, "fetch_with_playwright", return_value=sample_html) as pw,
            patch.object(sf, "fetch_with_curl_cffi", return_value=None),
            patch.object(sf, "fetch_with_requests", return_value=None),
            patch("sys.argv", ["scrape_footystats.py", "--refresh"]),
        ):
            sf.main()
        assert pw.call_count == 1
        sf.logger.handlers.clear()

        with (
            patch.object(sf, "fetch_html") as fetch_mock,
            patch("sys.argv", ["scrape_footystats.py"]),
        ):
            sf.main()
        fetch_mock.assert_not_called()
        sf.logger.handlers.clear()
