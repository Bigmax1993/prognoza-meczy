from openpyxl import load_workbook

import scrape_footystats as sf


def test_save_excel_creates_workbook(tmp_paths, sample_rows):
    out = tmp_paths["root"] / "out.xlsx"
    sf.save_excel(sample_rows, out)
    assert out.exists()

    wb = load_workbook(out)
    ws = wb.active
    assert ws.title == "Mecze"
    assert ws.cell(1, 1).value == "Data"
    assert ws.cell(2, 5).value == sample_rows[0]["Gospodarz"]
    assert ws.max_row == len(sample_rows) + 1


def test_save_excel_empty_rows(tmp_paths):
    out = tmp_paths["root"] / "empty.xlsx"
    sf.save_excel([], out)
    wb = load_workbook(out)
    assert wb.active["A1"].value == "Brak meczów"
