# -*- coding: utf-8 -*-
"""Testy regresyjne — kontrakt danych, golden fixture, ochrona przed regressjami."""
from __future__ import annotations

import csv
import hashlib
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

import scrape_footystats as sf
from tests.fixtures.sample_html import SAMPLE_MATCHES_HTML

GOLDEN_DIR = Path(__file__).parent / "fixtures" / "golden"
GOLDEN_CSV = GOLDEN_DIR / "sample_mecze.csv"

EXPECTED_COLUMNS = [
    "Data",
    "Kraj",
    "Liga",
    "Godzina",
    "Gospodarz",
    "Forma gospodarza",
    "Gość",
    "Forma gościa",
    "Kurs 1",
    "Kurs X",
    "Kurs 2",
    "Status",
    "Link",
]


def _row_fingerprint(rows: list[dict]) -> str:
    lines = ["|".join(str(r.get(c, "")) for c in EXPECTED_COLUMNS) for r in rows]
    return hashlib.sha256("\n".join(lines).encode("utf-8")).hexdigest()


@pytest.fixture(scope="module")
def parsed_fixture():
    return sf.parse_matches(SAMPLE_MATCHES_HTML)


class TestSchemaRegression:
    def test_row_columns_stable(self, parsed_fixture):
        _, rows = parsed_fixture
        assert list(rows[0].keys()) == EXPECTED_COLUMNS

    def test_excel_header_matches_schema(self, tmp_paths, parsed_fixture):
        _, rows = parsed_fixture
        out = tmp_paths["root"] / "schema.xlsx"
        sf.save_excel(rows, out)
        wb = load_workbook(out)
        headers = [wb.active.cell(1, i).value for i in range(1, len(EXPECTED_COLUMNS) + 1)]
        assert headers == EXPECTED_COLUMNS


class TestKnownMatchesRegression:
    """Zablokowane wartości z fixture HTML — wykrywają regresję parsera."""

    def test_first_match_locked_fields(self, parsed_fixture):
        match_date, rows = parsed_fixture
        assert match_date == "11/08"
        first = rows[0]
        assert first["Gospodarz"] == "Жальгирис"
        assert first["Gość"] == "Динамо Загреб"
        assert first["Godzina"] == "18:00"
        assert first["Forma gospodarza"] == "1.60"
        assert first["Forma gościa"] == "2.33"
        assert first["Kurs 1"] == "6.35"
        assert first["Kurs X"] == "4.62"
        assert first["Kurs 2"] == "1.39"
        assert first["Link"].endswith("/ru/europe/team-a-vs-team-b-h2h-stats")

    def test_match_count_floor_on_fixture(self, parsed_fixture):
        _, rows = parsed_fixture
        assert len(rows) == 3

    def test_england_league_name_not_duplicating_country(self, parsed_fixture):
        _, rows = parsed_fixture
        england = next(r for r in rows if r["Kraj"] == "Англия")
        assert england["Liga"] == "Кубок Англии"
        assert not england["Liga"].startswith("Англия")


class TestGoldenCsvRegression:
    def test_pipeline_matches_golden_csv(self, parsed_fixture):
        if not GOLDEN_CSV.exists():
            pytest.skip("brak golden CSV — uruchom: python tests/generate_golden.py")

        _, rows = parsed_fixture
        with GOLDEN_CSV.open(encoding="utf-8-sig", newline="") as f:
            expected = list(csv.DictReader(f))

        assert len(rows) == len(expected)
        assert _row_fingerprint(rows) == _row_fingerprint(expected)
        for exp, act in zip(expected, rows):
            for col in EXPECTED_COLUMNS:
                assert act[col] == exp[col], f"{col}: {act[col]!r} != {exp[col]!r}"


class TestLiveCacheRegression:
    """Ochrona przed regresją na realnym cache (jeśli obecny)."""

    def test_desktop_cache_minimum_and_schema(self):
        cache = Path(__file__).resolve().parents[1] / "cache" / "matches_cache.json"
        if not cache.exists():
            pytest.skip("brak cache/matches_cache.json")

        data = json.loads(cache.read_text(encoding="utf-8"))
        matches = data.get("matches") or []
        assert len(matches) >= 100
        assert data.get("match_date")
        sample = matches[0]
        for col in EXPECTED_COLUMNS:
            assert col in sample
        # godziny nie mogą tracić cyfr dziesiątek (historyczny bug scrapera tekstowego)
        times = {m.get("Godzina", "") for m in matches}
        assert any(t.startswith(("1", "2")) for t in times if t)


class TestEmptyInputRegression:
    def test_empty_html_returns_no_rows_but_date(self):
        match_date, rows = sf.parse_matches("<html><body></body></html>")
        assert rows == []
        assert match_date  # fallback DD/MM

    def test_empty_excel_placeholder(self, tmp_paths):
        out = tmp_paths["root"] / "empty.xlsx"
        sf.save_excel([], out)
        wb = load_workbook(out)
        assert wb.active["A1"].value == "Brak meczów"
