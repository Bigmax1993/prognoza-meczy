"""Testy całego pipeline: fetch → parse → cache → excel → logi."""

from __future__ import annotations

from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

import scrape_footystats as sf


def test_full_pipeline_from_html(tmp_paths, sample_html, configured_logger):
    """Pełny przepływ bez sieci: HTML → parse → cache JSON → Excel."""
    match_date, rows = sf.parse_matches(sample_html)
    assert len(rows) == 3

    cache_path = sf.save_matches_cache(match_date, rows, "fixture")
    sf.save_html_cache(sample_html, "fixture")

    out = tmp_paths["root"] / f"footystats_mecze_{match_date.replace('/', '-')}.xlsx"
    sf.save_excel(rows, out)

    # cache
    cached = sf.load_matches_cache(ttl_minutes=30)
    assert cached is not None
    assert cached[0] == match_date
    assert len(cached[1]) == 3
    assert cache_path.exists()
    assert tmp_paths["html_cache"].exists()

    # excel
    wb = load_workbook(out)
    ws = wb.active
    assert ws.max_row == 4  # header + 3 mecze
    assert ws.cell(2, 4).value == "18:00"

    # logi
    sf.logger.info("pipeline-ok")
    for h in sf.logger.handlers:
        h.flush()
    assert "pipeline-ok" in configured_logger.read_text(encoding="utf-8")


def test_main_uses_cache(tmp_paths, sample_rows, monkeypatch):
    """main() bez --refresh korzysta z cache (bez fetch)."""
    sf.save_matches_cache("11/08", sample_rows, "seed")
    sf.logger.handlers.clear()

    with (
        patch.object(sf, "fetch_html") as fetch_mock,
        patch("sys.argv", ["scrape_footystats.py"]),
    ):
        sf.main()

    fetch_mock.assert_not_called()
    excel_files = list(tmp_paths["root"].glob("footystats_mecze_*.xlsx"))
    assert len(excel_files) == 1
    sf.logger.handlers.clear()


def test_main_refresh_fetches(tmp_paths, sample_html, monkeypatch):
    """main(--refresh) pomija cache i woła fetch_html."""
    sf.save_matches_cache("11/08", [{"Data": "old"}], "old")
    sf.logger.handlers.clear()

    with (
        patch.object(sf, "fetch_html", return_value=(sample_html, "mock")) as fetch_mock,
        patch("sys.argv", ["scrape_footystats.py", "--refresh"]),
    ):
        sf.main()

    fetch_mock.assert_called_once()
    cached = sf.load_matches_cache(ttl_minutes=30)
    assert cached is not None
    assert len(cached[1]) == 3
    assert list(tmp_paths["root"].glob("footystats_mecze_*.xlsx"))
    sf.logger.handlers.clear()


def test_main_fetch_then_parse_excel_cache(tmp_paths, sample_html):
    """End-to-end main: mockowany fetch → parse → cache → excel."""
    sf.logger.handlers.clear()

    with (
            patch.object(sf, "fetch_with_playwright", return_value=sample_html),
            patch.object(sf, "fetch_with_curl_cffi", return_value=None),
            patch.object(sf, "fetch_with_requests", return_value=None),
            patch("sys.argv", ["scrape_footystats.py", "--refresh", "--log-level", "INFO"]),
        ):
            sf.main()

    assert tmp_paths["matches_cache"].exists()
    assert tmp_paths["html_cache"].exists()
    excel = next(tmp_paths["root"].glob("footystats_mecze_*.xlsx"))
    wb = load_workbook(excel)
    assert wb.active.max_row == 4

    log_files = list(tmp_paths["logs"].glob("scraper_*.log"))
    assert log_files
    log_text = log_files[0].read_text(encoding="utf-8")
    assert "Znaleziono 3 meczów" in log_text or "Znaleziono" in log_text
    sf.logger.handlers.clear()


def test_pipeline_second_run_hits_cache(tmp_paths, sample_html):
    """Pierwszy run scrapuje, drugi bierze cache."""
    sf.logger.handlers.clear()

    with (
        patch.object(sf, "fetch_with_playwright", return_value=sample_html) as pw,
        patch.object(sf, "fetch_with_curl_cffi", return_value=None),
        patch.object(sf, "fetch_with_requests", return_value=None),
        patch("sys.argv", ["scrape_footystats.py", "--refresh"]),
    ):
        sf.main()
    assert pw.call_count == 1
    sf.logger.handlers.clear()

    with (
        patch.object(sf, "fetch_html") as fetch_mock,
        patch("sys.argv", ["scrape_footystats.py"]),
    ):
        sf.main()
    fetch_mock.assert_not_called()
    sf.logger.handlers.clear()
