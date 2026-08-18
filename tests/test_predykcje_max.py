"""Testy modułu predykcje_max.py."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

import predykcje_max as pmax


@pytest.fixture
def form_history(tmp_path) -> Path:
    rows = []
    fixtures_2025 = [
        ("05/01/2025", "Arsenal", "Chelsea", "2:1", "так", 11, 4),
        ("12/01/2025", "Chelsea", "Arsenal", "0:2", "ні", 8, 2),
        ("19/01/2025", "Arsenal", "Liverpool", "3:1", "так", 12, 5),
        ("26/01/2025", "Liverpool", "Chelsea", "1:1", "так", 10, 3),
        ("02/02/2025", "Chelsea", "Liverpool", "2:0", "ні", 7, 2),
        ("09/02/2025", "Arsenal", "Chelsea", "1:0", "ні", 9, 3),
        ("16/02/2025", "Chelsea", "Arsenal", "1:3", "так", 14, 4),
        ("23/02/2025", "Liverpool", "Arsenal", "2:2", "так", 13, 5),
    ]
    fixtures_2024 = [
        ("05/01/2024", "Arsenal", "Chelsea", "1:1", "так", 10, 3),
        ("12/01/2024", "Chelsea", "Liverpool", "2:1", "так", 9, 4),
        ("19/01/2024", "Liverpool", "Arsenal", "0:1", "ні", 8, 2),
        ("26/01/2024", "Arsenal", "Liverpool", "2:0", "ні", 11, 3),
        ("02/02/2024", "Chelsea", "Arsenal", "1:2", "так", 12, 5),
    ]
    for d, h, a, score, btts, corners, cards in fixtures_2024 + fixtures_2025:
        rows.append(
            {
                "ліга": "Premier League",
                "дата": d,
                "господар": h,
                "гість": a,
                "результат": score,
                "оз": btts,
                "фоли_господар": 10,
                "фоли_гість": 12,
                "фоли": 22,
                "кутові_господар": corners // 2,
                "кутові_гість": corners - corners // 2,
                "кутові": corners,
                "жовті_картки_господар": cards // 2,
                "жовті_картки_гість": cards - cards // 2,
                "жовті_картки": cards,
                "удари_господар": 12,
                "удари_гість": 10,
                "удари": 22,
                "удари_в_площину_господар": 6,
                "удари_в_площину_гість": 4,
                "удари_в_площину": 10,
            }
        )
    rows.append(
        {
            "ліга": "Premier League",
            "дата": "01/03/2026",
            "господар": "Arsenal",
            "гість": "Chelsea",
            "результат": "2:1",
            "оз": "так",
            "фоли_господар": 9,
            "фоли_гість": 11,
            "фоли": 20,
            "кутові_господар": 6,
            "кутові_гість": 5,
            "кутові": 11,
            "жовті_картки_господар": 2,
            "жовті_картки_гість": 1,
            "жовті_картки": 3,
            "удари_господар": 14,
            "удари_гість": 9,
            "удари": 23,
            "удари_в_площину_господар": 7,
            "удари_в_площину_гість": 3,
            "удари_в_площину": 10,
        }
    )
    path = tmp_path / "form_data.csv"
    pd.DataFrame(rows).to_csv(path, index=False, encoding="utf-8-sig")
    return path


def test_poisson_probs_sum_to_100():
    ph, pd_, pa = pmax.poisson_1x2_probs(1.4, 1.1)
    assert pytest.approx(ph + pd_ + pa, abs=0.2) == 100


def test_has_league_stats_nordic(form_history):
    _, hist = pmax.base.load_2026_data(form_history)
    assert pmax.has_league_stats("Premier League", hist) == True
    nordic = hist.copy()
    nordic.loc[0, "ліга"] = "Allsvenskan"
    nordic["кутові"] = pd.NA
    assert pmax.has_league_stats("Allsvenskan", nordic) is False


def test_predict_match_max_keys(form_history):
    _, hist = pmax.base.load_2026_data(form_history)
    pred = pmax.predict_match_max(
        "Arsenal",
        "Chelsea",
        "Premier League",
        hist,
        as_of=pd.Timestamp("2026-03-01"),
    )
    required = {
        "przewidywany_zwyciezca",
        "prawdopodobienstwo_gospodarz",
        "predykcja_gole_2.5",
        "predykcja_rozne_9.5",
        "predykcja_zolte_3.5",
        "handicap_0",
        "pewnosc_score",
        "value_flag",
        "ma_statystyki",
    }
    assert required.issubset(pred.keys())
    assert pred["predykcja_gole_2.5"] in {"over", "under"}
    assert pred["przewidywane_btts"] in {"так", "ні"}


def test_build_predictions_max(form_history):
    df_2026, hist = pmax.base.load_2026_data(form_history)
    preds, feats = pmax.build_predictions_max(df_2026, hist)
    assert len(preds) == 1
    assert len(feats) == 1
    assert "predykcja_gole_1.5" in preds.columns
    assert "forma_3_h_gf" in feats.columns


def test_backtest_max(form_history):
    _, hist = pmax.base.load_2026_data(form_history)
    detail, summary = pmax.run_backtest_max(hist, year=2025)
    assert not detail.empty
    assert "hit_rate_1x2_pct" in summary["metryka"].values
    assert "hit_rate_gole_2_5_pct" in summary["metryka"].values


def test_benchmark(form_history):
    _, hist = pmax.base.load_2026_data(form_history)
    bench = pmax.run_benchmark(hist, year=2025)
    assert "metryka" in bench.columns
    assert "delta_pp" in bench.columns


def test_select_top_typy(form_history):
    df_2026, hist = pmax.base.load_2026_data(form_history)
    preds, _ = pmax.build_predictions_max(df_2026, hist)
    top = pmax.select_top_typy(preds, top_n=5)
    assert len(top) <= 5


def test_export_excel_max_sheets(form_history, tmp_path):
    df_2026, hist = pmax.base.load_2026_data(form_history)
    preds, feats = pmax.build_predictions_max(df_2026, hist)
    team_avg = pmax.base.compute_team_averages(hist)
    league_avg = pmax.base.compute_league_averages(df_2026)
    bt_d, bt_s = pmax.run_backtest_max(hist, year=2025)
    bench = pmax.run_benchmark(hist, year=2025)
    top = pmax.select_top_typy(preds)
    out = tmp_path / "predykcje_max_2026.xlsx"

    pmax.export_excel_max(
        df_2026,
        preds,
        feats,
        team_avg,
        league_avg,
        out,
        backtest_detail=bt_d,
        backtest_summary=bt_s,
        benchmark=bench,
        top_typy=top,
    )

    assert out.exists()
    wb = load_workbook(out)
    names = wb.sheetnames
    assert "Premier League" in names
    assert pmax.UA_SHEET_SUMMARY["top_typy"] in names
    assert pmax.UA_SHEET_SUMMARY["benchmark"] in names
    assert pmax.UA_SHEET_SUMMARY["audyt"] in names
    ws = wb["Premier League"]
    assert ws.tables
    headers = [c.value for c in ws[1]]
    assert "дата" in headers
    assert "прогноз_переможець" in headers
    assert "прогноз голів" in headers


def test_drop_empty_columns():
    df = pd.DataFrame(
        {
            "дата": ["01/01/2026", "02/01/2026"],
            "пусто": [None, None],
            "текст_пусто": ["", ""],
            "прогноз голів": ["так", "ні"],
        }
    )
    out = pmax.drop_empty_columns(df)
    assert list(out.columns) == ["дата", "прогноз голів"]


def test_translate_df_ua_over_under():
    df = pd.DataFrame({"predykcja_gole_2.5": ["over", "under"], "pewnosc": ["wysoka", "niska"]})
    ua = pmax.translate_df_ua(df)
    assert ua.iloc[0]["прогноз_голи_2_5"] == "більше"
    assert ua.iloc[1]["прогноз_голи_2_5"] == "менше"
    assert ua.iloc[0]["впевненість"] == "висока"


def test_league_calibration_fit(form_history):
    _, hist = pmax.base.load_2026_data(form_history)
    cal = pmax.LeagueCalibration.fit(hist)
    assert "Premier League" in cal.foul_lines or cal.foul_lines == {}
