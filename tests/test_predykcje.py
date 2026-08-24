"""Testy modułu predykcje.py (forma, O/U, backtest)."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

import predykcje as pred


@pytest.fixture
def sample_csv(tmp_path) -> Path:
    rows = [
        {
            "ліга": "Premier League",
            "дата": "01/01/2026",
            "господар": "Arsenal",
            "гість": "Chelsea",
            "оз": "так",
            "фоли": 20,
            "кутові": 10,
            "жовті_картки": 3,
            "удари": 25,
            "удари_в_площину": 12,
        },
        {
            "ліга": "Premier League",
            "дата": "02/01/2026",
            "господар": "Chelsea",
            "гість": "Arsenal",
            "оз": "ні",
            "фоли": 18,
            "кутові": 8,
            "жовті_картки": 2,
            "удари": 20,
            "удари_в_площину": 8,
        },
        {
            "ліга": "Premier League",
            "дата": "03/01/2026",
            "господар": "Arsenal",
            "гість": "Liverpool",
            "оз": "так",
            "фоли": 22,
            "кутові": 11,
            "жовті_картки": 4,
            "удари": 28,
            "удари_в_площину": 14,
        },
        {
            "ліга": "La Liga",
            "дата": "15/06/2025",
            "господар": "Real",
            "гість": "Barca",
            "оз": "так",
            "фоли": 15,
            "кутові": 9,
            "жовті_картки": 2,
            "удари": 22,
            "удари_в_площину": 10,
        },
    ]
    path = tmp_path / "aleks_ligi_stats.csv"
    pd.DataFrame(rows).to_csv(path, index=False, encoding="utf-8-sig")
    return path


@pytest.fixture
def form_history(tmp_path) -> Path:
    """Więcej meczów z wynikami — do formy i backtestu."""
    rows = []
    # 2025: 6 meczów Arsenal/Chelsea z wynikami
    fixtures_2025 = [
        ("05/01/2025", "Arsenal", "Chelsea", "2:1", "так", 11, 4),
        ("12/01/2025", "Chelsea", "Arsenal", "0:2", "ні", 8, 2),
        ("19/01/2025", "Arsenal", "Liverpool", "3:1", "так", 12, 5),
        ("26/01/2025", "Liverpool", "Chelsea", "1:1", "так", 10, 3),
        ("02/02/2025", "Chelsea", "Liverpool", "2:0", "ні", 7, 2),
        ("09/02/2025", "Arsenal", "Chelsea", "1:0", "ні", 9, 3),
        ("16/02/2025", "Chelsea", "Arsenal", "1:3", "так", 14, 4),
        ("23/02/2025", "Liverpool", "Arsenal", "2:2", "так", 13, 5),
    ]
    for d, h, a, score, btts, corners, cards in fixtures_2025:
        rows.append(
            {
                "ліга": "Premier League",
                "дата": d,
                "господар": h,
                "гість": a,
                "результат": score,
                "оз": btts,
                "фоли_господар": 10,
                "фоли_гість": 12,
                "фоли": 22,
                "кутові_господар": corners // 2,
                "кутові_гість": corners - corners // 2,
                "кутові": corners,
                "жовті_картки_господар": cards // 2,
                "жовті_картки_гість": cards - cards // 2,
                "жовті_картки": cards,
                "удари_господар": 12,
                "удари_гість": 10,
                "удари": 22,
                "удари_в_площину_господар": 6,
                "удари_в_площину_гість": 4,
                "удари_в_площину": 10,
            }
        )
    # 2026
    rows.append(
        {
            "ліга": "Premier League",
            "дата": "01/03/2026",
            "господар": "Arsenal",
            "гість": "Chelsea",
            "результат": "2:1",
            "оз": "так",
            "фоли_господар": 9,
            "фоли_гість": 11,
            "фоли": 20,
            "кутові_господар": 6,
            "кутові_гість": 5,
            "кутові": 11,
            "жовті_картки_господар": 2,
            "жовті_картки_гість": 1,
            "жовті_картки": 3,
            "удари_господар": 14,
            "удари_гість": 9,
            "удари": 23,
            "удари_в_площину_господар": 7,
            "удари_в_площину_гість": 3,
            "удари_в_площину": 10,
        }
    )
    path = tmp_path / "form_data.csv"
    pd.DataFrame(rows).to_csv(path, index=False, encoding="utf-8-sig")
    return path


def test_load_2026_filters_year(sample_csv):
    df_2026, df_hist = pred.load_2026_data(sample_csv)
    assert len(df_2026) == 3
    assert len(df_hist) == 4
    assert (df_2026["дата"].dt.year == 2026).all()


def test_filter_from_date_keeps_from_cutoff():
    df = pd.DataFrame(
        {
            "дата": pd.to_datetime(["2026-08-12", "2026-08-13", "2026-08-14"]),
            "ліга": ["Premier League"] * 3,
        }
    )
    out = pred.filter_from_date(df, pd.Timestamp("2026-08-13"))
    assert len(out) == 2
    assert out["дата"].min() == pd.Timestamp("2026-08-13")


def test_validate_from_date_accepts_window():
    df = pd.DataFrame({"дата": ["13/08/2026", "18/08/2026"]})
    pred.validate_from_date(df, pred.FROM_DATE, label="test")


def test_validate_from_date_rejects_earlier_matches():
    df = pd.DataFrame({"дата": ["12/08/2026", "13/08/2026"]})
    with pytest.raises(ValueError, match="przed 13/08/2026"):
        pred.validate_from_date(df, pred.FROM_DATE, label="test")


def test_compute_team_averages(sample_csv):
    _, df_hist = pred.load_2026_data(sample_csv)
    avg = pred.compute_team_averages(df_hist)
    assert "druzyna" in avg.columns
    assert "srednia_faule" in avg.columns
    arsenal = avg[avg["druzyna"] == "Arsenal"].iloc[0]
    assert arsenal["mecze"] == 3
    assert pytest.approx(arsenal["srednia_faule"], rel=0.01) == 10.0


def test_compute_league_averages(sample_csv):
    df_2026, _ = pred.load_2026_data(sample_csv)
    avg = pred.compute_league_averages(df_2026)
    assert len(avg) == 1
    assert avg.iloc[0]["ліга"] == "Premier League"
    assert avg.iloc[0]["mecze"] == 3


def test_predict_match_returns_required_keys(sample_csv):
    _, df_hist = pred.load_2026_data(sample_csv)
    team_avg = pred.compute_team_averages(df_hist)
    result = pred.predict_match("Arsenal", "Chelsea", "Premier League", df_hist, team_avg)
    required = {
        "przewidywany_zwyciezca",
        "prawdopodobienstwo_gospodarz",
        "prawdopodobienstwo_gosc",
        "prawdopodobienstwo_remis",
        "przewidywany_wynik",
        "pewnosc",
        "metoda",
        "predykcja_rozne",
        "predykcja_zolte",
        "status",
    }
    assert required.issubset(result.keys())
    assert ":" in result["przewidywany_wynik"]


def test_poisson_1x2_draw_depends_on_expected_goals():
    p_h, p_d, p_a = pred._poisson_1x2(1.2, 1.2)
    assert abs(p_h + p_d + p_a - 1.0) < 1e-9
    assert p_d > 0.20
    _, p_mismatch, _ = pred._poisson_1x2(2.6, 0.5)
    assert p_d > p_mismatch
    strong_h, _, weak_a = pred._poisson_1x2(2.4, 0.8)
    assert strong_h > weak_a


def test_compute_form_last_n(form_history):
    _, hist = pred.load_2026_data(form_history)
    as_of = pd.Timestamp("2026-03-01")
    form = pred.compute_form(hist, "Arsenal", as_of=as_of, n=5)
    assert form["mecze"] == 5
    assert form["srednia_goli_strzelonych"] is not None
    home_form = pred.compute_form(hist, "Arsenal", as_of=as_of, n=5, venue="home")
    assert home_form["mecze"] >= 1


def test_predict_includes_ou_and_form(form_history):
    df_2026, hist = pred.load_2026_data(form_history)
    pred_row = pred.predict_match(
        "Arsenal",
        "Chelsea",
        "Premier League",
        hist,
        as_of=pd.Timestamp("2026-03-01"),
    )
    assert pred_row["status"] == "ok"
    assert pred_row["predykcja_rozne"] in {"over", "under"}
    assert pred_row["predykcja_zolte"] in {"over", "under"}
    assert pred_row["forma_gospodarz_mecze"] >= 3
    assert pred_row["metoda"] in {pred.METODA_FORMA, pred.METODA_GOALS, pred.METODA_STATS}


def test_predict_fills_stat_fallbacks_when_history_has_no_corners(form_history):
    df_2026, hist = pred.load_2026_data(form_history)
    hist = hist.copy()
    for c in ("кутові", "кутові_господар", "кутові_гість", "жовті_картки", "жовті_картки_господар", "жовті_картки_гість"):
        if c in hist.columns:
            hist[c] = pd.NA
    row = pred.predict_match(
        "Arsenal",
        "Chelsea",
        "Premier League",
        hist,
        as_of=pd.Timestamp("2026-03-01"),
    )
    assert row["exp_rozne"] is not None
    assert row["exp_zolte"] is not None
    assert row["predykcja_rozne"] in {"over", "under"}
    assert row["srednia_rozne_gospodarz"] is not None


def test_build_predictions(sample_csv):
    df_2026, df_hist = pred.load_2026_data(sample_csv)
    preds = pred.build_predictions(df_2026, df_hist)
    assert len(preds) == 3
    assert "przewidywany_zwyciezca" in preds.columns
    assert "predykcja_rozne" in preds.columns


def test_backtest_2025(form_history):
    _, hist = pred.load_2026_data(form_history)
    detail, summary = pred.run_backtest(hist, year=2025)
    assert not detail.empty
    assert "hit_rate_1x2_pct" in summary["metryka"].values
    assert "trafiony_1x2" in detail.columns


def test_crosstab_liga_1x2(form_history):
    df_2026, df_hist = pred.load_2026_data(form_history)
    preds = pred.build_predictions(df_2026, df_hist)
    scored = pred._with_1x2(preds)
    pt = pred._crosstab_liga(scored, "_1x2", columns=["1", "X", "2"])
    assert not pt.empty
    assert "liga" in pt.columns
    assert set(["1", "X", "2"]).issubset(set(pt.columns))
    assert int(pt[["1", "X", "2"]].sum(axis=1).iloc[0]) == len(preds[preds["status"] != "pomin"])


def test_build_league_pivot_from_predictions(form_history):
    df_2026, df_hist = pred.load_2026_data(form_history)
    preds = pred.build_predictions(df_2026, df_hist)
    pivot = pred._build_league_pivot(preds, "Premier League")
    assert not pivot.empty
    assert "typowane_wygrane" in pivot.columns
    assert "btts_tak" in pivot.columns
    assert set(pivot["druzyna"]).issubset({"Arsenal", "Chelsea"})


def test_ukrainize_for_excel_keeps_team_names():
    df = pd.DataFrame(
        [
            {
                "ліга": "Premier League",
                "господар": "Arsenal",
                "гість": "Chelsea",
                "status": "ok",
                "przewidywany_zwyciezca": "Arsenal",
                "pewnosc": "wysoka",
                "predykcja_rozne": "over",
            },
            {
                "ліга": "Eliteserien",
                "господар": "Sarpsborg 08",
                "гість": "Sandefjord",
                "status": "pomin",
                "przewidywany_zwyciezca": "remis",
                "pewnosc": "niska",
                "predykcja_rozne": "under",
            },
        ]
    )
    out = pred.ukrainize_for_excel(df)
    assert list(out["господар"]) == ["Arsenal", "Sarpsborg 08"]
    assert "Прем'єр-ліга" in set(out["ліга"])
    assert "Елітесеріен" in set(out["ліга"])
    assert "статус" in out.columns
    assert list(out["статус"]) == ["ок", "пропуск"]
    assert out.iloc[0]["прогноз_переможець"] == "Arsenal"
    assert out.iloc[1]["прогноз_переможець"] == "нічия"
    assert list(out["прогноз_кутові"]) == ["більше", "менше"]


def test_ukrainize_renames_btts_column():
    df = pd.DataFrame({"ліга": ["Premier League"], "оз": ["так"]})
    out = pred.ukrainize_for_excel(df)
    assert "оз" not in out.columns
    assert pred.COL_BTTS_UA in out.columns
    assert list(out[pred.COL_BTTS_UA]) == ["так"]


def test_split_played_and_future_moves_unplayed():
    as_of = pd.Timestamp("2026-08-18")
    df = pd.DataFrame(
        {
            "ліга": ["Прем'єр-ліга", "Ла Ліга", "Серія А"],
            "дата": ["17/08/2026", "22/08/2026", "16/08/2026"],
            "господар": ["Arsenal", "Valencia", "Inter"],
            "гість": ["Chelsea", "Celta", "Monza"],
            "результат": ["2:1", None, "1:0"],
            "фоли": [20, None, 18],
        }
    )
    played, future = pred.split_played_and_future(df, as_of=as_of)
    assert list(played["господар"]) == ["Arsenal", "Inter"]
    assert list(future["господар"]) == ["Valencia"]
    assert list(future.columns) == list(df.columns)


def test_export_excel_creates_sheets(form_history, tmp_path):
    df_2026, df_hist = pred.load_2026_data(form_history)
    team_avg = pred.compute_team_averages(df_hist)
    league_avg = pred.compute_league_averages(df_2026)
    preds = pred.build_predictions(df_2026, df_hist)
    bt_d, bt_s = pred.run_backtest(df_hist, year=2025)
    out = tmp_path / "predykcje_2026.xlsx"

    pred.export_excel(
        df_2026,
        preds,
        team_avg,
        league_avg,
        out,
        backtest_detail=bt_d,
        backtest_summary=bt_s,
    )

    assert out.exists()
    wb = load_workbook(out)
    names = wb.sheetnames
    assert names == ["Матчі_2026", "Майбутні_матчі", "Прогнози"]

    def _assert_excel_table(sheet_name: str) -> None:
        ws = wb[sheet_name]
        assert len(ws.tables) >= 1, sheet_name
        tab = list(ws.tables.values())[0]
        style = tab.tableStyleInfo
        assert style is not None
        assert style.name == "TableStyleMedium2"
        assert style.showRowStripes is True

    _assert_excel_table("Матчі_2026")
    _assert_excel_table("Прогнози")
    mecze = pd.read_excel(out, sheet_name="Матчі_2026")
    assert "Прем'єр-ліга" in set(mecze["ліга"].astype(str))
    assert pred.COL_BTTS_UA in mecze.columns
    assert "оз" not in mecze.columns
    future_xl = pd.read_excel(out, sheet_name="Майбутні_матчі")
    assert list(future_xl.columns)[:4] == ["ліга", "дата", "господар", "гість"]
    preds_xl = pd.read_excel(out, sheet_name="Прогнози")
    assert "статус" in preds_xl.columns
    assert "прогноз_переможець" in preds_xl.columns
    assert "результат" not in preds_xl.columns
    assert "результат" in mecze.columns
    assert set(preds_xl["статус"].astype(str)).issubset({"ок", "пропуск"})


def test_export_excel_future_sheet_calendar_only(form_history, tmp_path):
    df_2026, df_hist = pred.load_2026_data(form_history)
    extra = pd.DataFrame(
        {
            "ліга": ["Premier League"],
            "дата": [pd.Timestamp("2026-08-22")],
            "господар": ["Everton"],
            "гість": ["Crystal Palace"],
            "результат": [pd.NA],
        }
    )
    extra = pred._to_numeric(extra)
    extra = pred._attach_goals(extra)
    df_2026 = pd.concat([df_2026, extra], ignore_index=True)
    team_avg = pred.compute_team_averages(df_hist)
    league_avg = pred.compute_league_averages(df_2026)
    preds = pred.build_predictions(df_2026, df_hist)
    out = tmp_path / "predykcje_2026.xlsx"
    pred.export_excel(
        df_2026,
        preds,
        team_avg,
        league_avg,
        out,
        as_of=pd.Timestamp("2026-08-18"),
    )
    mecze = pd.read_excel(out, sheet_name="Матчі_2026")
    future = pd.read_excel(out, sheet_name="Майбутні_матчі")
    assert "Everton" not in set(mecze["господар"].astype(str))
    assert list(future.columns) == ["ліга", "дата", "господар", "гість"]
    assert list(future["господар"]) == ["Everton"]
    assert list(future["гість"]) == ["Crystal Palace"]
    assert "результат" not in future.columns
    preds_xl = pd.read_excel(out, sheet_name="Прогнози")
    assert "результат" not in preds_xl.columns
    eve = preds_xl[preds_xl["господар"].astype(str) == "Everton"].iloc[0]
    assert str(eve["прогноз_рахунок"]).strip() not in {"", "—", "nan"}
    assert ":" in str(eve["прогноз_рахунок"])


def test_export_excel_played_sheet_from_cutoff_keeps_recent_ft(form_history, tmp_path):
    df_2026, df_hist = pred.load_2026_data(form_history)
    extra = pd.DataFrame(
        {
            "ліга": ["Premier League", "Premier League"],
            "дата": [pd.Timestamp("2026-01-10"), pd.Timestamp("2026-08-22")],
            "господар": ["Liverpool", "Everton"],
            "гість": ["Chelsea", "Leeds"],
            "результат": ["2:0", "1:0"],
        }
    )
    extra = pred._to_numeric(extra)
    extra = pred._attach_goals(extra)
    df_all = pd.concat([df_2026, extra], ignore_index=True)
    window = pred.filter_from_date(df_all, pred.FROM_DATE)
    team_avg = pred.compute_team_averages(df_hist)
    league_avg = pred.compute_league_averages(window)
    preds = pred.build_predictions(window, df_hist)
    out = tmp_path / "predykcje_2026.xlsx"
    pred.export_excel(
        window,
        preds,
        team_avg,
        league_avg,
        out,
        from_date=pred.FROM_DATE,
        as_of=pd.Timestamp("2026-08-24"),
    )
    mecze = pd.read_excel(out, sheet_name="Матчі_2026")
    homes = set(mecze["господар"].astype(str))
    assert "Liverpool" not in homes
    assert "Everton" in homes
    dates = pd.to_datetime(mecze["дата"], dayfirst=True)
    assert dates.min() >= pred.FROM_DATE
