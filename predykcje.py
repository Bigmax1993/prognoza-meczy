# -*- coding: utf-8 -*-
"""Predykcje meczów + Excel (rok 2026) + backtest 2025.

Heurystyka:
  - forma ostatnich N meczów (ogólna + home/away) mieszana z sezonem
  - jeśli jest результат → średnie bramek (model_wynikow / forma)
  - inaczej: удари_в_площину (heurystyka_statystyk)
  - Over/Under: кутові 9.5, жовті_картки 3.5
  - BTTS z % оз / wyniku
  - backtest na mecze 2025 (predykcja tylko z historii sprzed daty)

Uruchomienie:
  python predykcje.py
  python predykcje.py --od 13/08/2026
"""
from __future__ import annotations

import argparse
import logging
import math
import re
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrData, StrRef, StrVal
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

from preview_pandas import ALEKS_XLSX, read_aleks_table, resolve_aleks_path
from team_names import canonical_name, canonicalize_teams, known_teams_by_league

ROOT = Path(__file__).resolve().parent
OUT_XLSX = ROOT / "predykcje_2026.xlsx"
FROM_DATE = pd.Timestamp("2026-08-13")

logger = logging.getLogger("predykcje")

COL_LIGA = "ліга"
COL_DATA = "дата"
COL_HOME = "господар"
COL_AWAY = "гість"
COL_BTTS = "оз"
COL_BTTS_UA = "Чи обидві забили?"
COL_RESULT = "результат"

STAT_COLS = ["фоли", "кутові", "жовті_картки", "удари", "удари_в_площину"]
STAT_LABELS = {
    "фоли": "faule",
    "кутові": "rozne",
    "жовті_картки": "zolte_kartki",
    "удари": "strzaly",
    "удари_в_площину": "strzaly_w_stwor",
}

HOME_ADVANTAGE = 1.12
SOT_TO_GOALS = 0.28
FORM_N = 5
FORM_WEIGHT = 0.70  # 70% forma, 30% sezon
SEASON_WEIGHT = 0.30
MIN_MATCHES = 3
CORNERS_LINE = 9.5
CARDS_LINE = 3.5
METODA_STATS = "heurystyka_statystyk"
METODA_GOALS = "model_wynikow"
METODA_FORMA = "forma_5"

SHEET_MECZE = "Матчі_2026"
SHEET_FUTURE = "Майбутні_матчі"
SHEET_PRED = "Прогнози"
FUTURE_COLS = [COL_LIGA, COL_DATA, COL_HOME, COL_AWAY]

LEAGUE_UA = {
    "Premier League": "Прем'єр-ліга",
    "La Liga": "Ла Ліга",
    "Serie A": "Серія А",
    "Bundesliga": "Бундесліга",
    "Bundesliga 2": "Друга Бундесліга",
    "Eredivisie": "Ередивізі",
    "Allsvenskan": "Аллсвенскан",
    "Eliteserien": "Елітесеріен",
    "Super League": "Суперліга",
    "Championship": "Чемпіоншип",
}

EXCEL_UA_COL = {
    "status": "статус",
    "powod": "причина",
    "przewidywany_zwyciezca": "прогноз_переможець",
    "prawdopodobienstwo_gospodarz": "ймовірність_господар",
    "prawdopodobienstwo_remis": "ймовірність_нічия",
    "prawdopodobienstwo_gosc": "ймовірність_гість",
    "przewidywany_wynik": "прогноз_рахунок",
    "оз": COL_BTTS_UA,
    "przewidywane_btts": "прогноз голів",
    "przewidywane_btts_pct": "прогноз_оз_відсоток",
    "exp_rozne": "очікувані_кутові",
    "exp_zolte": "очікувані_жовті",
    "linia_rozne": "лінія_кутові",
    "linia_zolte": "лінія_жовті",
    "forma_gospodarz_mecze": "форма_господар_матчі",
    "forma_gosc_mecze": "форма_гість_матчі",
    "metoda": "метод",
    "pewnosc": "впевненість",
    "predykcja_rozne": "прогноз_кутові",
    "predykcja_zolte": "прогноз_жовті",
    "forma_gospodarz_goli": "форма_господар_голи",
    "forma_gosc_goli": "форма_гість_голи",
    "min_meczy_historii": "мін_матчів_історії",
    "srednia_faule_gospodarz": "середн_фоли_господар",
    "srednia_faule_gosc": "середн_фоли_гість",
    "srednia_rozne_gospodarz": "середн_кутові_господар",
    "srednia_rozne_gosc": "середн_кутові_гість",
    "srednia_zolte_gospodarz": "середн_жовті_господар",
    "srednia_zolte_gosc": "середн_жовті_гість",
    "srednia_strzaly_gospodarz": "середн_удари_господар",
    "srednia_strzaly_gosc": "середн_удари_гість",
    "srednia_strzaly_stwor_gospodarz": "середн_удари_в_площину_господар",
    "srednia_strzaly_stwor_gosc": "середн_удари_в_площину_гість",
    "srednia_goli_gospodarz": "середн_голи_господар",
    "srednia_goli_gosc": "середн_голи_гість",
}

EXCEL_UA_VAL = {
    "ok": "ок",
    "pomin": "пропуск",
    "za_malo_danych": "замало_даних",
    "remis": "нічия",
    "over": "більше",
    "under": "менше",
    "wysoka": "висока",
    "srednia": "середня",
    "niska": "низька",
    "forma_5": "форма_5",
    "heurystyka_statystyk": "евристика_статистик",
    "model_wynikow": "модель_рахунку",
    "tak": "так",
    "nie": "ні",
    "—": "—",
}

DEFAULT_STATS = {
    "фоли": 22.0,
    "кутові": 10.0,
    "жовті_картки": 3.5,
    "удари": 22.0,
    "удари_в_площину": 8.0,
}


def _safe_print(text: str) -> None:
    try:
        if hasattr(sys.stdout, "reconfigure"):
            try:
                sys.stdout.reconfigure(errors="replace")
            except Exception:
                pass
        print(text)
    except UnicodeEncodeError:
        enc = sys.stdout.encoding or "utf-8"
        print(text.encode(enc, errors="replace").decode(enc, errors="replace"))


def ukrainize_for_excel(df: pd.DataFrame, *, fill_blank: str | None = None) -> pd.DataFrame:
    """Nagłówki i wartości po ukraińsku; nazwy drużyn bez zmian."""
    if df is None or df.empty:
        return df
    out = df.copy()
    team_cols = {COL_HOME, COL_AWAY, "przewidywany_zwyciezca", "прогноз_переможець"}
    if COL_LIGA in out.columns:
        out[COL_LIGA] = out[COL_LIGA].map(lambda x: LEAGUE_UA.get(str(x), x) if pd.notna(x) else x)
    for col in list(out.columns):
        if col in team_cols:
            out[col] = out[col].map(
                lambda x: EXCEL_UA_VAL["remis"] if str(x) == "remis" else x
            )
            continue
        if col == COL_LIGA:
            continue
        if pd.api.types.is_numeric_dtype(out[col]) or pd.api.types.is_datetime64_any_dtype(out[col]):
            continue
        out[col] = out[col].map(
            lambda x: EXCEL_UA_VAL[str(x)] if pd.notna(x) and str(x) in EXCEL_UA_VAL else x
        )
    if fill_blank:
        for col in out.columns:
            if col in team_cols:
                out[col] = out[col].map(
                    lambda x, fb=fill_blank: fb if _excel_blank(x) else x
                )
                continue
            if pd.api.types.is_numeric_dtype(out[col]):
                continue
            out[col] = out[col].map(lambda x, fb=fill_blank: fb if _excel_blank(x) else x)
    out = out.rename(columns={k: v for k, v in EXCEL_UA_COL.items() if k in out.columns})
    return out


def _excel_blank(v: object) -> bool:
    if v is None:
        return True
    try:
        if pd.isna(v):
            return True
    except (TypeError, ValueError):
        pass
    return str(v).strip() in {"", "nan", "None", "<NA>", "—"}


def split_played_and_future(
    df: pd.DataFrame,
    *,
    as_of: pd.Timestamp | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Rozegrane (wynik albo data < dziś) vs przyszłe bez wyniku."""
    if df is None or df.empty:
        empty = df.copy() if df is not None else pd.DataFrame(columns=FUTURE_COLS)
        return empty, pd.DataFrame(columns=FUTURE_COLS)
    cut = (as_of or pd.Timestamp.now()).normalize()
    if COL_DATA in df.columns:
        dates = pd.to_datetime(df[COL_DATA], dayfirst=True, errors="coerce")
        past = dates.dt.normalize() < cut
    else:
        past = pd.Series(False, index=df.index)
    if COL_RESULT in df.columns:
        has_score = df[COL_RESULT].map(lambda x: not _excel_blank(x))
    else:
        has_score = pd.Series(False, index=df.index)
    played_mask = has_score | past.fillna(False)
    played = df.loc[played_mask].copy()
    future = df.loc[~played_mask].copy()
    return played, future


def parse_score(value: object) -> tuple[int, int] | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    text = str(value).strip()
    m = re.match(r"^(\d+)\s*[:\-]\s*(\d+)$", text)
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))


def _to_numeric(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in STAT_COLS:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce")
        for sfx in ("_господар", "_гість"):
            c = f"{col}{sfx}"
            if c in out.columns:
                out[c] = pd.to_numeric(out[c], errors="coerce")
    return out


def _attach_goals(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if COL_RESULT not in out.columns:
        out["_hg"] = pd.NA
        out["_ag"] = pd.NA
        return out
    parsed = out[COL_RESULT].map(parse_score)
    out["_hg"] = parsed.map(lambda x: x[0] if x else pd.NA)
    out["_ag"] = parsed.map(lambda x: x[1] if x else pd.NA)
    return out


def parse_od_date(text: str) -> pd.Timestamp:
    ts = pd.to_datetime(str(text).strip(), dayfirst=True, errors="coerce")
    if pd.isna(ts):
        raise argparse.ArgumentTypeError(f"Niepoprawna data: {text}")
    return pd.Timestamp(ts).normalize()


def filter_from_date(df: pd.DataFrame, from_date: pd.Timestamp | None) -> pd.DataFrame:
    """Zostawia mecze od from_date (włącznie)."""
    if from_date is None or df.empty or COL_DATA not in df.columns:
        return df
    dates = pd.to_datetime(df[COL_DATA], dayfirst=True, errors="coerce")
    return df.loc[dates >= pd.Timestamp(from_date)].copy()


def validate_from_date(
    df: pd.DataFrame,
    from_date: pd.Timestamp,
    *,
    label: str = "dane",
) -> None:
    """Walidacja: żaden mecz nie może być sprzed from_date."""
    if df.empty or COL_DATA not in df.columns:
        raise ValueError(f"Walidacja ({label}): brak kolumny daty albo puste dane.")
    dates = pd.to_datetime(df[COL_DATA], dayfirst=True, errors="coerce")
    if dates.isna().any():
        n_bad = int(dates.isna().sum())
        raise ValueError(f"Walidacja ({label}): {n_bad} wierszy bez poprawnej daty.")
    too_early = dates < pd.Timestamp(from_date)
    if too_early.any():
        n = int(too_early.sum())
        earliest = pd.Timestamp(dates.min()).strftime("%d/%m/%Y")
        cutoff = pd.Timestamp(from_date).strftime("%d/%m/%Y")
        raise ValueError(
            f"Walidacja ({label}): {n} meczow przed {cutoff} (najwczesniejszy: {earliest})."
        )


def load_data(
    path: Path | None = None,
    *,
    year: int | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Wczytuje dane. Zwraca (df_year lub cały, df_full)."""
    src = Path(path) if path is not None else resolve_aleks_path()
    if src is None:
        raise FileNotFoundError(f"Brak pliku: {ALEKS_XLSX}")

    df = read_aleks_table(src)
    df[COL_DATA] = pd.to_datetime(df[COL_DATA], format="%d/%m/%Y", errors="coerce")
    df = _to_numeric(df)
    df = _attach_goals(df)
    df = canonicalize_teams(df)
    df = df.dropna(subset=[COL_DATA]).sort_values(COL_DATA).reset_index(drop=True)

    if year is None:
        return df.copy(), df
    df_y = df[df[COL_DATA].dt.year == year].copy()
    return df_y, df


def load_2026_data(
    csv_path: Path | None = None,
    *,
    include_history: bool = True,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Kompatybilność: (df_2026, historia)."""
    df_2026, full = load_data(csv_path, year=2026)
    history = full if include_history else df_2026
    if df_2026.empty:
        logger.warning("Brak wierszy z 2026 (przed: %s)", len(full))
    return df_2026, history


def _same_team(a: object, b: object) -> bool:
    return canonical_name(str(a or "")) == canonical_name(str(b or ""))


def _team_match_stat(row: pd.Series, team: str, base: str) -> float | None:
    hcol, acol = f"{base}_господар", f"{base}_гість"
    if _same_team(row[COL_HOME], team) and hcol in row.index and pd.notna(row[hcol]):
        return float(row[hcol])
    if _same_team(row[COL_AWAY], team) and acol in row.index and pd.notna(row[acol]):
        return float(row[acol])
    if base in row.index and pd.notna(row[base]):
        return float(row[base]) / 2
    return None


def _team_goals(row: pd.Series, team: str) -> tuple[float, float] | None:
    if pd.isna(row.get("_hg")) or pd.isna(row.get("_ag")):
        return None
    if _same_team(row[COL_HOME], team):
        return float(row["_hg"]), float(row["_ag"])
    if _same_team(row[COL_AWAY], team):
        return float(row["_ag"]), float(row["_hg"])
    return None


def _mean(vals: list[float]) -> float | None:
    return sum(vals) / len(vals) if vals else None


def _poisson_pmf(k: int, lam: float) -> float:
    lam = max(0.05, float(lam))
    return math.exp(-lam) * (lam ** k) / math.factorial(k)


def _poisson_1x2(exp_home: float, exp_away: float, max_g: int = 8) -> tuple[float, float, float]:
    """P(1), P(X), P(2) z niezależnego Poissona oczekiwanych goli."""
    p_home = p_draw = p_away = 0.0
    for hg in range(max_g + 1):
        ph = _poisson_pmf(hg, exp_home)
        for ag in range(max_g + 1):
            p = ph * _poisson_pmf(ag, exp_away)
            if hg > ag:
                p_home += p
            elif hg == ag:
                p_draw += p
            else:
                p_away += p
    total = p_home + p_draw + p_away or 1.0
    return p_home / total, p_draw / total, p_away / total


class TeamStatsIndex:
    """Szybki indeks meczów per drużyna (forma bez skanowania całego DF)."""

    def __init__(self, df: pd.DataFrame) -> None:
        self._by_team: dict[str, list[dict[str, Any]]] = {}
        data = df.sort_values(COL_DATA)
        for _, m in data.iterrows():
            date = pd.Timestamp(m[COL_DATA])
            for team, venue in ((m[COL_HOME], "home"), (m[COL_AWAY], "away")):
                g = _team_goals(m, team)
                rec: dict[str, Any] = {
                    "date": date,
                    "venue": venue,
                    "gf": g[0] if g else None,
                    "ga": g[1] if g else None,
                    "btts": str(m.get(COL_BTTS, "")).strip().casefold(),
                }
                for base in STAT_COLS:
                    rec[base] = _team_match_stat(m, team, base)
                self._by_team.setdefault(canonical_name(str(team)), []).append(rec)

    def form(
        self,
        team: str,
        *,
        as_of: pd.Timestamp | None = None,
        n: int = FORM_N,
        venue: str | None = None,
    ) -> dict[str, Any]:
        recs = self._by_team.get(canonical_name(str(team)), [])
        if not recs:
            recs = self._by_team.get(str(team), [])
        if as_of is not None:
            recs = [r for r in recs if r["date"] < as_of]
        if venue in {"home", "away"}:
            recs = [r for r in recs if r["venue"] == venue]
        recs = recs[-n:]
        gf = [r["gf"] for r in recs if _is_num(r["gf"])]
        ga = [r["ga"] for r in recs if _is_num(r["ga"])]
        stats = {c: [r[c] for r in recs if _is_num(r[c])] for c in STAT_COLS}
        btts_yes = 0
        for r in recs:
            if r["btts"] in {"yes", "tak", "так"}:
                btts_yes += 1
            elif r["gf"] is not None and r["ga"] is not None and r["gf"] > 0 and r["ga"] > 0:
                btts_yes += 1
        out: dict[str, Any] = {
            "mecze": len(recs),
            "srednia_goli_strzelonych": _mean(gf),
            "srednia_goli_straconych": _mean(ga),
            "btts_yes_pct": (btts_yes / len(recs) * 100) if recs else None,
        }
        for base in STAT_COLS:
            out[f"srednia_{STAT_LABELS[base]}"] = _mean(stats[base])
        return out


def get_team_index(df: pd.DataFrame) -> TeamStatsIndex:
    return TeamStatsIndex(df)


def team_matches_before(
    df: pd.DataFrame,
    team: str,
    *,
    as_of: pd.Timestamp | None = None,
    venue: str | None = None,
) -> pd.DataFrame:
    """Mecze drużyny przed datą as_of (wyłącznie). venue: home|away|None."""
    key = canonical_name(team)
    home_ok = df[COL_HOME].map(lambda x: canonical_name(str(x))) == key
    away_ok = df[COL_AWAY].map(lambda x: canonical_name(str(x))) == key
    mask = home_ok | away_ok
    if venue == "home":
        mask = home_ok
    elif venue == "away":
        mask = away_ok
    sub = df.loc[mask]
    if as_of is not None:
        sub = sub[sub[COL_DATA] < as_of]
    return sub.sort_values(COL_DATA)


def compute_form(
    df: pd.DataFrame,
    team: str,
    *,
    as_of: pd.Timestamp | None = None,
    n: int = FORM_N,
    venue: str | None = None,
    index: TeamStatsIndex | None = None,
) -> dict[str, Any]:
    """Średnie z ostatnich n meczów (ew. tylko home/away)."""
    if index is not None:
        return index.form(team, as_of=as_of, n=n, venue=venue)
    sub = team_matches_before(df, team, as_of=as_of, venue=venue).tail(n)
    gf: list[float] = []
    ga: list[float] = []
    stats: dict[str, list[float]] = {c: [] for c in STAT_COLS}
    btts_yes = 0
    for _, m in sub.iterrows():
        g = _team_goals(m, team)
        if g:
            gf.append(g[0])
            ga.append(g[1])
        for base in STAT_COLS:
            v = _team_match_stat(m, team, base)
            if v is not None:
                stats[base].append(v)
        bval = str(m.get(COL_BTTS, "")).strip().casefold()
        if bval in {"yes", "tak", "так"} or (g and g[0] > 0 and g[1] > 0):
            btts_yes += 1

    out: dict[str, Any] = {
        "mecze": len(sub),
        "srednia_goli_strzelonych": _mean(gf),
        "srednia_goli_straconych": _mean(ga),
        "btts_yes_pct": (btts_yes / len(sub) * 100) if len(sub) else None,
    }
    for base in STAT_COLS:
        out[f"srednia_{STAT_LABELS[base]}"] = _mean(stats[base])
    return out


def compute_team_averages(
    df: pd.DataFrame,
    *,
    as_of: pd.Timestamp | None = None,
) -> pd.DataFrame:
    """Średnie sezonowe per drużyna (opcjonalnie tylko mecze przed as_of)."""
    data = df if as_of is None else df[df[COL_DATA] < as_of]
    rows: list[dict[str, Any]] = []
    teams = pd.unique(
        pd.concat([data[COL_HOME], data[COL_AWAY]]).map(lambda x: canonical_name(str(x)))
    )

    for team in teams:
        sub = team_matches_before(data, team)
        if sub.empty:
            continue
        entry: dict[str, Any] = {
            "druzyna": team,
            "mecze": len(sub),
            COL_LIGA: sub[COL_LIGA].mode().iloc[0] if not sub[COL_LIGA].mode().empty else "",
        }
        for col in STAT_COLS:
            vals = [_team_match_stat(m, team, col) for _, m in sub.iterrows()]
            vals_f = [v for v in vals if v is not None]
            entry[f"srednia_{STAT_LABELS[col]}"] = _mean(vals_f)

        gf: list[float] = []
        ga: list[float] = []
        for _, m in sub.iterrows():
            g = _team_goals(m, team)
            if g:
                gf.append(g[0])
                ga.append(g[1])
        entry["srednia_goli_strzelonych"] = _mean(gf)
        entry["srednia_goli_straconych"] = _mean(ga)
        entry["mecze_z_wynikiem"] = len(gf)

        btts = sub[COL_BTTS].astype(str).str.strip().str.casefold()
        entry["btts_yes_pct"] = btts.isin(["yes", "tak", "так"]).mean() * 100
        if ((btts == "") | (btts == "nan")).all() and gf:
            both = sum(1 for a, b in zip(gf, ga) if a > 0 and b > 0)
            entry["btts_yes_pct"] = both / len(gf) * 100
        rows.append(entry)

    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows).sort_values(["druzyna"]).reset_index(drop=True)


def compute_league_averages(df: pd.DataFrame) -> pd.DataFrame:
    agg: dict[str, Any] = {"mecze": (COL_LIGA, "count")}
    for col in STAT_COLS:
        agg[f"srednia_{STAT_LABELS[col]}"] = (col, "mean")
    out = df.groupby(COL_LIGA, as_index=False).agg(**agg)
    btts = (
        df.assign(
            _btts=df[COL_BTTS]
            .astype(str)
            .str.strip()
            .str.casefold()
            .isin(["yes", "tak", "так"])
        )
        .groupby(COL_LIGA)["_btts"]
        .mean()
        .mul(100)
        .reset_index(name="btts_yes_pct")
    )
    return out.merge(btts, on=COL_LIGA).sort_values(COL_LIGA).reset_index(drop=True)


def _is_num(v: object) -> bool:
    if v is None:
        return False
    try:
        if pd.isna(v):
            return False
    except (TypeError, ValueError):
        pass
    try:
        float(v)
    except (TypeError, ValueError):
        return False
    return True


def _col_mean(df: pd.DataFrame, col: str, league: str | None = None) -> float | None:
    if df is None or df.empty or col not in df.columns:
        return None
    data = df
    if league and COL_LIGA in df.columns:
        lig = df[df[COL_LIGA] == league]
        if not lig.empty and pd.to_numeric(lig[col], errors="coerce").notna().any():
            data = lig
    s = pd.to_numeric(data[col], errors="coerce")
    return float(s.mean()) if s.notna().any() else None


def _stat_or_default(
    form: dict[str, Any],
    season: dict[str, Any],
    key: str,
    *,
    base_col: str,
    pool: pd.DataFrame,
    league: str,
) -> float:
    blended = _blend(form.get(key), season.get(key))
    if _is_num(blended):
        return float(blended)
    lig = _col_mean(pool, base_col, league)
    if lig is not None:
        return lig
    glob = _col_mean(pool, base_col)
    if glob is not None:
        return glob
    return float(DEFAULT_STATS.get(base_col, 0.0))


def _blend(form_val: float | None, season_val: float | None) -> float | None:
    f_ok, s_ok = _is_num(form_val), _is_num(season_val)
    if f_ok and s_ok:
        return FORM_WEIGHT * float(form_val) + SEASON_WEIGHT * float(season_val)
    if f_ok:
        return float(form_val)
    if s_ok:
        return float(season_val)
    return None


def _confidence(gap: float, home_n: int, away_n: int) -> str:
    min_n = min(home_n, away_n)
    if min_n < MIN_MATCHES:
        return "niska"
    if gap >= 0.25 and min_n >= 5:
        return "wysoka"
    if gap >= 0.12 and min_n >= 3:
        return "srednia"
    return "niska"


def _ou_label(expected: float | None, line: float) -> str:
    if expected is None:
        return ""
    return "over" if expected > line else "under"


def predict_match(
    home: str,
    away: str,
    league: str,
    df_history: pd.DataFrame,
    team_avg: pd.DataFrame | None = None,
    *,
    as_of: pd.Timestamp | None = None,
    index: TeamStatsIndex | None = None,
) -> dict[str, Any]:
    """Predykcja z formą N meczów + O/U rożne/kartki."""
    hist = df_history
    pool = hist if as_of is None else hist[hist[COL_DATA] < as_of]
    if team_avg is None:
        team_avg = compute_team_averages(
            hist if as_of is None else hist[hist[COL_DATA] < as_of]
        )
    idx = index or get_team_index(df_history)
    home = canonical_name(str(home))
    away = canonical_name(str(away))

    def season_row(team: str) -> dict[str, Any]:
        if team_avg.empty:
            return {}
        key = canonical_name(team)
        hit = team_avg[team_avg["druzyna"].map(lambda x: canonical_name(str(x))) == key]
        return hit.iloc[0].to_dict() if not hit.empty else {}

    h_s = season_row(home)
    a_s = season_row(away)
    h_form = compute_form(hist, home, as_of=as_of, n=FORM_N, index=idx)
    a_form = compute_form(hist, away, as_of=as_of, n=FORM_N, index=idx)
    h_home = compute_form(hist, home, as_of=as_of, n=FORM_N, venue="home", index=idx)
    a_away = compute_form(hist, away, as_of=as_of, n=FORM_N, venue="away", index=idx)

    def pick(form: dict, season: dict, key: str) -> float | None:
        return _blend(form.get(key), season.get(key))

    # preferuj formę home/away do bramek gdy jest
    home_gf = _blend(
        h_home.get("srednia_goli_strzelonych") or h_form.get("srednia_goli_strzelonych"),
        h_s.get("srednia_goli_strzelonych"),
    )
    home_ga = _blend(
        h_home.get("srednia_goli_straconych") or h_form.get("srednia_goli_straconych"),
        h_s.get("srednia_goli_straconych"),
    )
    away_gf = _blend(
        a_away.get("srednia_goli_strzelonych") or a_form.get("srednia_goli_strzelonych"),
        a_s.get("srednia_goli_strzelonych"),
    )
    away_ga = _blend(
        a_away.get("srednia_goli_straconych") or a_form.get("srednia_goli_straconych"),
        a_s.get("srednia_goli_straconych"),
    )

    home_gf = home_gf if home_gf is not None else 1.3
    home_ga = home_ga if home_ga is not None else 1.3
    away_gf = away_gf if away_gf is not None else 1.3
    away_ga = away_ga if away_ga is not None else 1.3

    home_n = int(h_form.get("mecze") or h_s.get("mecze") or 0)
    away_n = int(a_form.get("mecze") or a_s.get("mecze") or 0)
    home_n_goals = int(h_s.get("mecze_z_wynikiem") or 0)
    away_n_goals = int(a_s.get("mecze_z_wynikiem") or 0)

    skip = home_n < MIN_MATCHES or away_n < MIN_MATCHES

    use_goals = (
        home_gf is not None
        and away_gf is not None
        and home_ga is not None
        and away_ga is not None
        and (home_n_goals >= 2 or (h_form.get("mecze") or 0) >= 2)
        and (away_n_goals >= 2 or (a_form.get("mecze") or 0) >= 2)
    )

    if use_goals:
        metoda = METODA_FORMA if (h_form.get("mecze") or 0) >= 3 and (a_form.get("mecze") or 0) >= 3 else METODA_GOALS
        exp_home = max(0.2, (home_gf + away_ga) / 2 * HOME_ADVANTAGE)
        exp_away = max(0.2, (away_gf + home_ga) / 2)
        home_attack = exp_home
        away_attack = exp_away
    else:
        metoda = METODA_STATS
        league_sub = hist[hist[COL_LIGA] == league]
        league_sot = league_sub["удари_в_площину"].mean() if not league_sub.empty else 8.0
        if pd.isna(league_sot) or league_sot <= 0:
            league_sot = 8.0

        def sot(form: dict, season: dict) -> float:
            v = pick(form, season, "srednia_strzaly_w_stwor")
            return float(v) if v and v > 0 else float(league_sot)

        home_attack = sot(h_form, h_s) * HOME_ADVANTAGE
        away_attack = sot(a_form, a_s)
        exp_home = max(0.2, home_attack * SOT_TO_GOALS)
        exp_away = max(0.2, away_attack * SOT_TO_GOALS)

    p_home, p_draw, p_away = _poisson_1x2(exp_home, exp_away)

    if skip:
        winner = "—"
        pewnosc = "pomin"
        status = "pomin"
        powod = "za_malo_danych"
    else:
        status = "ok"
        powod = "—"
        if p_home >= p_away and p_home >= p_draw:
            winner = home
        elif p_away >= p_home and p_away >= p_draw:
            winner = away
        else:
            winner = "remis"
        gap = max(p_home, p_away, p_draw) - sorted([p_home, p_away, p_draw])[1]
        pewnosc = _confidence(gap, home_n, away_n)

    home_goals = max(0, round(exp_home))
    away_goals = max(0, round(exp_away))
    if winner == home and home_goals <= away_goals:
        home_goals = away_goals + 1
    elif winner == away and away_goals <= home_goals:
        away_goals = home_goals + 1
    elif winner == "remis" and home_goals != away_goals:
        home_goals = away_goals = max(home_goals, away_goals)

    h_btts = pick(h_form, h_s, "btts_yes_pct") or 50.0
    a_btts = pick(a_form, a_s, "btts_yes_pct") or 50.0
    btts_pct = (h_btts + a_btts) / 2
    btts_pred = "так" if btts_pct >= 50 else "ні"

    # Over/Under
    h_corners = _stat_or_default(h_form, h_s, "srednia_rozne", base_col="кутові", pool=pool, league=league)
    a_corners = _stat_or_default(a_form, a_s, "srednia_rozne", base_col="кутові", pool=pool, league=league)
    exp_corners = h_corners + a_corners
    h_cards = _stat_or_default(h_form, h_s, "srednia_zolte_kartki", base_col="жовті_картки", pool=pool, league=league)
    a_cards = _stat_or_default(a_form, a_s, "srednia_zolte_kartki", base_col="жовті_картки", pool=pool, league=league)
    exp_cards = h_cards + a_cards

    def rnd(v: float | None) -> float | None:
        if not _is_num(v):
            return None
        return round(float(v), 2)

    return {
        "status": status,
        "powod": powod,
        "przewidywany_zwyciezca": winner,
        "prawdopodobienstwo_gospodarz": round(p_home * 100, 1),
        "prawdopodobienstwo_gosc": round(p_away * 100, 1),
        "prawdopodobienstwo_remis": round(p_draw * 100, 1),
        "przewidywany_wynik": f"{home_goals}:{away_goals}",
        "przewidywane_btts": btts_pred,
        "przewidywane_btts_pct": round(btts_pct, 1),
        "exp_rozne": rnd(exp_corners),
        "predykcja_rozne": _ou_label(exp_corners, CORNERS_LINE),
        "linia_rozne": CORNERS_LINE,
        "exp_zolte": rnd(exp_cards),
        "predykcja_zolte": _ou_label(exp_cards, CARDS_LINE),
        "linia_zolte": CARDS_LINE,
        "forma_gospodarz_mecze": h_form.get("mecze", 0),
        "forma_gosc_mecze": a_form.get("mecze", 0),
        "forma_gospodarz_goli": rnd(h_form.get("srednia_goli_strzelonych") or home_gf),
        "forma_gosc_goli": rnd(a_form.get("srednia_goli_strzelonych") or away_gf),
        "metoda": metoda,
        "pewnosc": pewnosc,
        "min_meczy_historii": min(home_n, away_n),
        "srednia_faule_gospodarz": rnd(_stat_or_default(h_form, h_s, "srednia_faule", base_col="фоли", pool=pool, league=league)),
        "srednia_faule_gosc": rnd(_stat_or_default(a_form, a_s, "srednia_faule", base_col="фоли", pool=pool, league=league)),
        "srednia_rozne_gospodarz": rnd(h_corners),
        "srednia_rozne_gosc": rnd(a_corners),
        "srednia_zolte_gospodarz": rnd(h_cards),
        "srednia_zolte_gosc": rnd(a_cards),
        "srednia_strzaly_gospodarz": rnd(_stat_or_default(h_form, h_s, "srednia_strzaly", base_col="удари", pool=pool, league=league)),
        "srednia_strzaly_gosc": rnd(_stat_or_default(a_form, a_s, "srednia_strzaly", base_col="удари", pool=pool, league=league)),
        "srednia_strzaly_stwor_gospodarz": rnd(_stat_or_default(h_form, h_s, "srednia_strzaly_w_stwor", base_col="удари_в_площину", pool=pool, league=league)),
        "srednia_strzaly_stwor_gosc": rnd(_stat_or_default(a_form, a_s, "srednia_strzaly_w_stwor", base_col="удари_в_площину", pool=pool, league=league)),
        "srednia_goli_gospodarz": rnd(home_gf),
        "srednia_goli_gosc": rnd(away_gf),
    }


def build_predictions(
    df_target: pd.DataFrame,
    df_history: pd.DataFrame,
    *,
    use_as_of: bool = True,
) -> pd.DataFrame:
    """Predykcje dla meczów w df_target (domyślnie z historią sprzed daty meczu)."""
    known = known_teams_by_league(df_history)
    hist = canonicalize_teams(df_history, known)
    target = canonicalize_teams(df_target, known)
    rows: list[dict[str, Any]] = []
    # średnie sezonowe raz (przed pierwszym meczem targetu) — forma jest as_of
    if use_as_of and not target.empty:
        cut = pd.Timestamp(target[COL_DATA].min())
        season_avg = compute_team_averages(hist, as_of=cut)
    else:
        season_avg = compute_team_averages(hist)
    index = get_team_index(hist)

    for _, m in target.iterrows():
        as_of = pd.Timestamp(m[COL_DATA]) if use_as_of else None
        pred = predict_match(
            m[COL_HOME],
            m[COL_AWAY],
            m[COL_LIGA],
            hist,
            team_avg=season_avg,
            as_of=as_of,
            index=index,
        )
        actual = ""
        if COL_RESULT in m.index and pd.notna(m[COL_RESULT]):
            actual = str(m[COL_RESULT]).strip()
        rows.append(
            {
                COL_LIGA: m[COL_LIGA],
                COL_DATA: m[COL_DATA],
                COL_HOME: m[COL_HOME],
                COL_AWAY: m[COL_AWAY],
                COL_RESULT: actual,
                **pred,
            }
        )

    out = pd.DataFrame(rows)
    if not out.empty:
        out[COL_DATA] = pd.to_datetime(out[COL_DATA]).dt.strftime("%d/%m/%Y")
    return out


def _actual_winner(row: pd.Series) -> str:
    score = parse_score(row.get(COL_RESULT))
    if not score:
        return ""
    hg, ag = score
    if hg > ag:
        return str(row[COL_HOME])
    if ag > hg:
        return str(row[COL_AWAY])
    return "remis"


def _actual_btts(row: pd.Series) -> str:
    b = str(row.get(COL_BTTS, "")).strip().casefold()
    if b in {"yes", "tak", "так"}:
        return "так"
    if b in {"no", "nie", "ні"}:
        return "ні"
    score = parse_score(row.get(COL_RESULT))
    if score:
        return "так" if score[0] > 0 and score[1] > 0 else "ні"
    return ""


def run_backtest(
    df_history: pd.DataFrame,
    *,
    year: int = 2025,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Backtest: predykcje na mecze `year` tylko z historii sprzed daty."""
    target = df_history[df_history[COL_DATA].dt.year == year].copy()
    if target.empty:
        return pd.DataFrame(), pd.DataFrame()

    preds = build_predictions(target, df_history, use_as_of=True)
    # dołącz actuals z target (preds ma datę jako string)
    detail = preds.copy()
    detail["actual_zwyciezca"] = [
        _actual_winner(r) for _, r in target.reset_index(drop=True).iterrows()
    ]
    detail["actual_btts"] = [
        _actual_btts(r) for _, r in target.reset_index(drop=True).iterrows()
    ]
    detail["actual_rozne"] = pd.to_numeric(target["кутові"].values, errors="coerce") if "кутові" in target.columns else pd.NA
    detail["actual_zolte"] = (
        pd.to_numeric(target["жовті_картки"].values, errors="coerce")
        if "жовті_картки" in target.columns
        else pd.NA
    )

    scored = detail[detail["status"] == "ok"].copy()
    scored = scored[scored["actual_zwyciezca"].astype(str).str.len() > 0]

    def hit_rate(mask: pd.Series, correct: pd.Series) -> float | None:
        sub = correct[mask]
        return round(float(sub.mean() * 100), 1) if len(sub) else None

    if scored.empty:
        return detail, pd.DataFrame([{"metryka": "brak_danych", "wartosc": 0}])

    ok_1x2 = scored["przewidywany_zwyciezca"] == scored["actual_zwyciezca"]
    ok_btts = scored["przewidywane_btts"] == scored["actual_btts"]

    has_corners = scored["actual_rozne"].notna() & (scored["predykcja_rozne"].astype(str) != "")
    actual_corners_ou = scored["actual_rozne"].map(
        lambda x: "over" if pd.notna(x) and float(x) > CORNERS_LINE else ("under" if pd.notna(x) else "")
    )
    ok_corners = scored["predykcja_rozne"] == actual_corners_ou

    has_cards = scored["actual_zolte"].notna() & (scored["predykcja_zolte"].astype(str) != "")
    actual_cards_ou = scored["actual_zolte"].map(
        lambda x: "over" if pd.notna(x) and float(x) > CARDS_LINE else ("under" if pd.notna(x) else "")
    )
    ok_cards = scored["predykcja_zolte"] == actual_cards_ou

    summary_rows = [
        {"metryka": "rok", "wartosc": year},
        {"metryka": "mecze_razem", "wartosc": len(detail)},
        {"metryka": "mecze_ocenione", "wartosc": len(scored)},
        {"metryka": "pominiete", "wartosc": int((detail["status"] == "pomin").sum())},
        {"metryka": "hit_rate_1x2_pct", "wartosc": round(float(ok_1x2.mean() * 100), 1)},
        {"metryka": "hit_rate_btts_pct", "wartosc": round(float(ok_btts.mean() * 100), 1)},
        {
            "metryka": "hit_rate_rozne_9_5_pct",
            "wartosc": hit_rate(has_corners, ok_corners) if has_corners.any() else None,
        },
        {
            "metryka": "hit_rate_zolte_3_5_pct",
            "wartosc": hit_rate(has_cards, ok_cards) if has_cards.any() else None,
        },
    ]

    # per liga
    for liga, grp in scored.groupby(COL_LIGA):
        summary_rows.append(
            {
                "metryka": f"hit_1x2_{liga}",
                "wartosc": round(float((grp["przewidywany_zwyciezca"] == grp["actual_zwyciezca"]).mean() * 100), 1),
            }
        )

    detail["trafiony_1x2"] = (detail["przewidywany_zwyciezca"] == detail["actual_zwyciezca"]).map(
        {True: "так", False: "ні"}
    )
    return detail, pd.DataFrame(summary_rows)


def _table_display_name(base: str, used: set[str]) -> str:
    """Unikalna nazwa tabeli Excel (litery/cyfry/podkreślenie, nie zaczyna się od cyfry)."""
    clean = re.sub(r"[^\w]", "_", base, flags=re.UNICODE)
    clean = re.sub(r"_+", "_", clean).strip("_")
    if not clean or clean[0].isdigit():
        clean = "T_" + (clean or "data")
    name = clean[:25]
    i = 1
    while name in used:
        suffix = f"_{i}"
        name = clean[: 25 - len(suffix)] + suffix
        i += 1
    used.add(name)
    return name


def _excel_table_style() -> TableStyleInfo:
    return TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )


def _add_excel_table(
    ws,
    *,
    name_base: str,
    table_names: set[str],
    min_row: int = 1,
    min_col: int = 1,
    max_row: int | None = None,
    max_col: int | None = None,
) -> None:
    """Owija zakres w tabelę Excel (filtr, wiersz nagłówka, naprzemienne wiersze)."""
    max_row = max_row if max_row is not None else ws.max_row
    max_col = max_col if max_col is not None else ws.max_column
    if max_row < min_row + 1 or max_col < min_col:
        return
    ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    tab = Table(displayName=_table_display_name(name_base, table_names), ref=ref)
    tab.tableStyleInfo = _excel_table_style()
    ws.add_table(tab)


def _autosize_columns(ws, *, max_width: int = 22) -> None:
    for col_idx in range(1, (ws.max_column or 1) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width, 16)


def _write_df_sheet(
    wb: Workbook,
    title: str,
    df: pd.DataFrame,
    *,
    table_names: set[str],
) -> None:
    ws = wb.create_sheet(title=title[:31])
    export_df = df.drop(columns=[c for c in ("_hg", "_ag") if c in df.columns], errors="ignore")
    if export_df.empty:
        if list(export_df.columns):
            ws.append([str(c) for c in export_df.columns])
        else:
            ws.append(["немає_даних"])
        return
    for r in dataframe_to_rows(export_df, index=False, header=True):
        ws.append([_excel_value(v) for v in r])
    _add_excel_table(ws, name_base=title, table_names=table_names)
    _autosize_columns(ws)


def _excel_value(v: Any) -> Any:
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, pd.Timestamp):
        return v.strftime("%d/%m/%Y")
    if hasattr(v, "item") and not isinstance(v, (bytes, str)):
        try:
            return v.item()
        except (ValueError, AttributeError):
            pass
    return v


def _scored_predictions(predictions: pd.DataFrame) -> pd.DataFrame:
    if predictions.empty:
        return predictions
    out = predictions.copy()
    if "status" in out.columns:
        out = out[out["status"] != "pomin"]
    return out


def _with_1x2(predictions: pd.DataFrame) -> pd.DataFrame:
    """1 = gospodarz, X = remis, 2 = gość."""
    out = predictions.copy()

    def lab(row: pd.Series) -> str:
        w = str(row.get("przewidywany_zwyciezca") or "").strip()
        if not w:
            return ""
        if w == "remis":
            return "X"
        home = str(row.get(COL_HOME) or "").strip()
        away = str(row.get(COL_AWAY) or "").strip()
        if w == home:
            return "1"
        if w == away:
            return "2"
        return ""

    out["_1x2"] = out.apply(lab, axis=1)
    return out


def _crosstab_liga(
    df: pd.DataFrame,
    col: str,
    *,
    columns: list[str] | None = None,
) -> pd.DataFrame:
    """Tabela przestawna: liga × wartości kolumny (liczba meczów)."""
    work = _scored_predictions(df)
    if work.empty or COL_LIGA not in work.columns or col not in work.columns:
        return pd.DataFrame()
    labels = work[col].astype(str).str.strip()
    work = work.loc[labels.ne("") & labels.ne("nan")].copy()
    if work.empty:
        return pd.DataFrame()
    work[col] = work[col].astype(str).str.strip()
    pt = pd.crosstab(work[COL_LIGA], work[col])
    if columns:
        pt = pt.reindex(columns=columns, fill_value=0)
    return pt.reset_index().rename(columns={COL_LIGA: "liga"})


def _with_row_totals(pt: pd.DataFrame) -> pd.DataFrame:
    if pt.empty:
        return pt
    out = pt.copy()
    num_cols = [c for c in out.columns if c != "liga"]
    if not num_cols:
        return out
    out["Razem"] = out[num_cols].sum(axis=1)
    total = {c: int(out[c].sum()) for c in num_cols}
    total["liga"] = "Razem"
    total["Razem"] = int(out["Razem"].sum())
    return pd.concat([out, pd.DataFrame([total])], ignore_index=True)


def _prediction_pivot_frames(predictions: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """Tabele przestawne z arkusza Predykcje (tylko status ok)."""
    scored = _with_1x2(_scored_predictions(predictions))
    return {
        "1x2": _with_row_totals(_crosstab_liga(scored, "_1x2", columns=["1", "X", "2"])),
        "btts": _with_row_totals(_crosstab_liga(scored, "przewidywane_btts", columns=["так", "ні"])),
        "rozne": _with_row_totals(_crosstab_liga(scored, "predykcja_rozne", columns=["over", "under"])),
        "zolte": _with_row_totals(_crosstab_liga(scored, "predykcja_zolte", columns=["over", "under"])),
        "pewnosc": _with_row_totals(
            _crosstab_liga(scored, "pewnosc", columns=["wysoka", "srednia", "niska"])
        ),
        "_scored": scored,
    }


def _write_block(
    ws,
    start_row: int,
    title: str,
    df: pd.DataFrame,
    *,
    table_names: set[str] | None = None,
    table_key: str = "",
) -> tuple[int, int, int, int]:
    """Wpisuje tytuł + tabelę Excel. Zwraca (next_row, header_row, last_data_row, ncols)."""
    ws.cell(row=start_row, column=1, value=title)
    if df is None or df.empty:
        ws.cell(row=start_row + 1, column=1, value="brak danych")
        return start_row + 3, 0, 0, 0
    header_row = start_row + 1
    for c, name in enumerate(df.columns, start=1):
        ws.cell(row=header_row, column=c, value=str(name))
    for r_i, vals in enumerate(df.itertuples(index=False, name=None), start=header_row + 1):
        for c, v in enumerate(vals, start=1):
            ws.cell(row=r_i, column=c, value=_excel_value(v))
    last = header_row + len(df)
    ncols = int(len(df.columns))
    if table_names is not None:
        _add_excel_table(
            ws,
            name_base=table_key or title,
            table_names=table_names,
            min_row=header_row,
            max_row=last,
            max_col=ncols,
        )
    return last + 2, header_row, last, ncols


def _str_ref(formula: str, labels: list[str] | None = None) -> StrRef:
    cache = None
    if labels:
        pts = [StrVal(idx=i, v=str(v)) for i, v in enumerate(labels)]
        cache = StrData(ptCount=len(labels), pt=pts)
    return StrRef(f=formula, strCache=cache)


def _force_str_categories(chart, formula: str, labels: list[str] | None = None) -> None:
    """openpyxl zapisuje osie jako numRef — Excel wtedy często nie rysuje wykresu."""
    for ser in chart.series:
        ser.cat = AxDataSource(strRef=_str_ref(formula, labels))


def _add_pie_chart(
    ws,
    *,
    title: str,
    header_row: int,
    last_row: int,
    anchor: str,
    cat_labels: list[str] | None = None,
) -> None:
    if header_row < 1 or last_row <= header_row:
        return
    pie = PieChart()
    pie.title = title
    labels = Reference(ws, min_col=1, min_row=header_row + 1, max_row=last_row)
    data = Reference(ws, min_col=2, min_row=header_row, max_row=last_row)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    formula = f"{quote_sheetname(ws.title)}!$A${header_row + 1}:$A${last_row}"
    _force_str_categories(pie, formula, cat_labels)
    pie.width = 14
    pie.height = 10
    ws.add_chart(pie, anchor)


def _counts_frame(series: pd.Series, *, col: str, order: list[str] | None = None) -> pd.DataFrame:
    s = series.astype(str).str.strip()
    vc = s.value_counts()
    if order:
        vc = vc.reindex(order, fill_value=0)
        vc = vc[vc > 0]
    if vc.empty:
        return pd.DataFrame()
    return pd.DataFrame({col: vc.index.astype(str), "mecze": vc.values})


def _add_bar_chart(
    ws,
    *,
    title: str,
    header_row: int,
    last_row: int,
    ncols: int,
    anchor: str,
    cat_labels: list[str] | None = None,
) -> None:
    if header_row < 1 or last_row <= header_row or ncols < 2:
        return
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = title
    chart.y_axis.title = "mecze"
    chart.style = 10
    chart.x_axis.axPos = "b"
    chart.y_axis.axPos = "l"
    chart.x_axis.crosses = "min"
    chart.y_axis.crosses = "min"
    data = Reference(ws, min_col=2, min_row=header_row, max_col=ncols, max_row=last_row)
    cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    col = get_column_letter(1)
    formula = f"{quote_sheetname(ws.title)}!${col}${header_row + 1}:${col}${last_row}"
    _force_str_categories(chart, formula, cat_labels)
    chart.width = 18
    chart.height = 10
    if chart.legend is not None:
        chart.legend.position = "b"
    ws.add_chart(chart, anchor)


def _build_league_pivot(predictions: pd.DataFrame, league: str) -> pd.DataFrame:
    """Pivot per drużyna: typowane wygrane / BTTS / O-U (nie średnie gospodarza)."""
    sub = _scored_predictions(predictions)
    sub = sub[sub[COL_LIGA] == league].copy() if COL_LIGA in sub.columns else sub
    if sub.empty:
        return pd.DataFrame()
    teams = sorted(
        {str(t) for t in pd.concat([sub[COL_HOME], sub[COL_AWAY]], ignore_index=True).dropna().unique()}
    )
    rows: list[dict[str, Any]] = []
    for team in teams:
        played = sub[(sub[COL_HOME] == team) | (sub[COL_AWAY] == team)]
        n = len(played)
        if n == 0:
            continue
        wins = int((played["przewidywany_zwyciezca"] == team).sum()) if "przewidywany_zwyciezca" in played.columns else 0
        draws = int((played["przewidywany_zwyciezca"] == "remis").sum()) if "przewidywany_zwyciezca" in played.columns else 0
        btts_tak = int((played["przewidywane_btts"] == "так").sum()) if "przewidywane_btts" in played.columns else 0
        rozne_over = int((played["predykcja_rozne"] == "over").sum()) if "predykcja_rozne" in played.columns else 0
        zolte_over = int((played["predykcja_zolte"] == "over").sum()) if "predykcja_zolte" in played.columns else 0
        pewnosc_wysoka = int((played["pewnosc"] == "wysoka").sum()) if "pewnosc" in played.columns else 0
        rows.append(
            {
                COL_LIGA: league,
                "druzyna": team,
                "mecze": n,
                "typowane_wygrane": wins,
                "typowane_remisy": draws,
                "btts_tak": btts_tak,
                "rozne_over": rozne_over,
                "zolte_over": zolte_over,
                "pewnosc_wysoka": pewnosc_wysoka,
            }
        )
    if not rows:
        return pd.DataFrame()
    return (
        pd.DataFrame(rows)
        .sort_values(["typowane_wygrane", "druzyna"], ascending=[False, True])
        .reset_index(drop=True)
    )


def _league_bar_data(predictions: pd.DataFrame, league: str) -> pd.DataFrame:
    sub = _scored_predictions(predictions)
    sub = sub[sub[COL_LIGA] == league] if COL_LIGA in sub.columns else sub
    if sub.empty or "przewidywany_zwyciezca" not in sub.columns:
        return pd.DataFrame(columns=["druzyna", "typowane_wygrane"])
    wins: dict[str, int] = {}
    for _, r in sub.iterrows():
        w = str(r.get("przewidywany_zwyciezca") or "").strip()
        if not w or w == "remis":
            continue
        wins[w] = wins.get(w, 0) + 1
    if not wins:
        return pd.DataFrame(columns=["druzyna", "typowane_wygrane"])
    return pd.DataFrame(
        [{"druzyna": k, "typowane_wygrane": v} for k, v in sorted(wins.items(), key=lambda x: (-x[1], x[0]))]
    )


def _cat_labels(df: pd.DataFrame) -> list[str]:
    if df is None or df.empty:
        return []
    return [str(v) for v in df.iloc[:, 0].tolist()]


def _write_pivot_sheet(
    wb: Workbook,
    predictions: pd.DataFrame,
    *,
    table_names: set[str],
) -> dict[str, pd.DataFrame]:
    """Arkusz Pivot — tabele przestawne z Predykcji."""
    frames = _prediction_pivot_frames(predictions)
    ws = wb.create_sheet(title="Pivot")
    ws["A1"] = "Tabele przestawne z arkusza Predykcje (status ok)"
    ws["A2"] = "1 = gospodarz, X = remis, 2 = gość  |  Wykresy: arkusz Wykresy"
    row = 4
    blocks = (
        ("1X2", frames["1x2"], "Pivot_1X2"),
        ("BTTS", frames["btts"], "Pivot_BTTS"),
        (f"Rożne linia {CORNERS_LINE}", frames["rozne"], "Pivot_Rozne"),
        (f"Żółte kartki linia {CARDS_LINE}", frames["zolte"], "Pivot_Zolte"),
        ("Pewność typu", frames["pewnosc"], "Pivot_Pewnosc"),
    )
    for title, frame, key in blocks:
        row, _, _, _ = _write_block(ws, row, title, frame, table_names=table_names, table_key=key)
    ws.column_dimensions["A"].width = 28
    for col in ("B", "C", "D", "E"):
        ws.column_dimensions[col].width = 14
    return frames


def _chart_from_pivot(pt: pd.DataFrame) -> pd.DataFrame:
    """Pivot bez wiersza/kolumny Razem — pod wykres."""
    if pt is None or pt.empty:
        return pd.DataFrame()
    out = pt.copy()
    if "liga" in out.columns:
        out = out[out["liga"].astype(str) != "Razem"]
    if "Razem" in out.columns:
        out = out.drop(columns=["Razem"])
    return out.reset_index(drop=True)


def _write_charts_sheet(
    wb: Workbook,
    predictions: pd.DataFrame,
    frames: dict[str, pd.DataFrame],
    *,
    table_names: set[str],
) -> None:
    """Wykresy z tabel przestawnych arkusza Predykcje."""
    ws = wb.create_sheet(title="Wykresy")
    ws["A1"] = "Wykresy z tabeli przestawnej Predykcje. Otwórz w Microsoft Excel."
    ws.column_dimensions["A"].width = 22

    scored = frames.get("_scored")
    if scored is None:
        scored = _with_1x2(_scored_predictions(predictions))

    chart_frames = (
        ("1X2 wg ligi", _chart_from_pivot(frames.get("1x2")), "1X2", "Chart_1X2"),
        ("BTTS wg ligi", _chart_from_pivot(frames.get("btts")), "BTTS", "Chart_BTTS"),
        (f"Rożne {CORNERS_LINE} wg ligi", _chart_from_pivot(frames.get("rozne")), "Rozne", "Chart_Rozne"),
        (f"Żółte {CARDS_LINE} wg ligi", _chart_from_pivot(frames.get("zolte")), "Zolte", "Chart_Zolte"),
    )

    chart_row = 3
    for title, frame, short, key in chart_frames:
        start = chart_row
        next_row, header_row, last_row, ncols = _write_block(
            ws, start, title, frame, table_names=table_names, table_key=key
        )
        _add_bar_chart(
            ws,
            title=short,
            header_row=header_row,
            last_row=last_row,
            ncols=ncols,
            anchor=f"E{start}",
            cat_labels=_cat_labels(frame),
        )
        chart_row = max(next_row, start + 20)

    pew_df = _counts_frame(
        scored.get("pewnosc", pd.Series(dtype=str)) if not scored.empty else pd.Series(dtype=str),
        col="pewnosc",
        order=["wysoka", "srednia", "niska"],
    )
    pew_start = chart_row
    next_row, header_row, last_row, _ = _write_block(
        ws, pew_start, "Pewność typów (łącznie)", pew_df, table_names=table_names, table_key="Chart_Pewnosc"
    )
    _add_pie_chart(
        ws,
        title="Pewność",
        header_row=header_row,
        last_row=last_row,
        anchor=f"E{pew_start}",
        cat_labels=_cat_labels(pew_df),
    )


def export_excel(
    df_2026: pd.DataFrame,
    predictions: pd.DataFrame,
    team_avg: pd.DataFrame,
    league_avg: pd.DataFrame,
    path: Path | None = None,
    *,
    monthly: pd.DataFrame | None = None,
    backtest_detail: pd.DataFrame | None = None,
    backtest_summary: pd.DataFrame | None = None,
    from_date: pd.Timestamp | None = None,
    as_of: pd.Timestamp | None = None,
) -> Path:
    out_path = path or OUT_XLSX
    if from_date is not None:
        validate_from_date(df_2026, from_date, label="Mecze")
        validate_from_date(predictions, from_date, label="Predykcje")
    wb = Workbook()
    wb.remove(wb.active)
    table_names: set[str] = set()

    df_raw = ukrainize_for_excel(df_2026.copy())
    df_raw[COL_DATA] = pd.to_datetime(df_raw[COL_DATA], dayfirst=True).dt.strftime("%d/%m/%Y")
    played, future = split_played_and_future(df_raw, as_of=as_of)
    future = future.reindex(columns=FUTURE_COLS)
    preds_out = predictions.drop(columns=[COL_RESULT], errors="ignore")
    preds_ua = ukrainize_for_excel(preds_out, fill_blank="—")

    _write_df_sheet(wb, SHEET_MECZE, played, table_names=table_names)
    _write_df_sheet(wb, SHEET_FUTURE, future, table_names=table_names)
    _write_df_sheet(wb, SHEET_PRED, preds_ua, table_names=table_names)

    try:
        wb.save(out_path)
    except PermissionError:
        alt = out_path.with_name(f"{out_path.stem}_wypelnione{out_path.suffix}")
        wb.save(alt)
        logger.warning("Plik otwarty w Excelu — zapisano: %s", alt)
        return alt
    logger.info("Zapisano Excel: %s", out_path)
    return out_path


def _fill_played_from_json_and_api(
    df_2026: pd.DataFrame,
    df_history: pd.DataFrame,
    df_year: pd.DataFrame,
    *,
    from_date: pd.Timestamp,
    require_complete: bool,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Zawsze: skan braków → JSON → API (gdy są klucze). `--fill-missing` nie pozwala na puste pola."""
    import fill_missing as fill
    import upcoming as up

    if require_complete and not fill.has_keys():
        raise SystemExit(
            "Brak SERPER_API_KEY / ANTHROPIC_API_KEY — --fill-missing wymaga kluczy."
        )
    live = fill.has_keys()
    if not live:
        _safe_print("Braki: zapisuje JSON bez API (brak SERPER_API_KEY / ANTHROPIC_API_KEY)")
    try:
        df_2026, inv = fill.verify_and_fill(
            df_2026, live=live, max_rounds=3, from_date=from_date
        )
    except fill.ClaudeAuthError as exc:
        logger.warning("%s", exc)
        _safe_print(str(exc))
        inv = fill.load_missing_json(fill.STATS_JSON)
        df_2026 = fill.apply_inventory(df_2026, inv)
        df_2026 = fill.complete_row_totals(df_2026)
        raise SystemExit(str(exc)) from exc
    s = (inv or {}).get("summary") or {}
    v = (inv or {}).get("verification") or {}
    _safe_print(
        f"Braki (JSON {fill.STATS_JSON.name}): "
        f"luk={s.get('gaps', 0)} pending={s.get('pending', 0)} "
        f"wypelnione={s.get('filled', 0)}"
    )
    if v:
        _safe_print(
            f"Weryfikacja: rundy={v.get('rounds', 0)} "
            f"zostalo_meczow={v.get('remaining_matches', 0)} "
            f"zostalo_pol={v.get('remaining_fields', 0)}"
        )
    if inv:
        df_history = fill.apply_inventory(df_history, inv)
        df_history = fill.complete_row_totals(df_history)
    df_year = up.upsert_matches(df_year, df_2026)
    return df_2026, df_history, df_year


def _verify_exported_stats(
    path: Path,
    *,
    from_date: pd.Timestamp,
    require_complete: bool,
) -> tuple[pd.DataFrame | None, pd.DataFrame | None, bool]:
    """Po zapisie Excela: puste komórki → JSON ponownie → API na resztę."""
    import fill_missing as fill

    live = fill.has_keys()
    try:
        mecze_x, preds_x, vrep = fill.verify_exported_workbook(
            path,
            live=live,
            from_date=from_date,
        )
    except fill.ClaudeAuthError as exc:
        logger.warning("%s", exc)
        _safe_print(str(exc))
        raise SystemExit(str(exc)) from exc
    _safe_print(
        f"Weryfikacja Excela: puste {vrep['empty_before']['fields']} pol → "
        f"z JSON {vrep['filled_from_json']}, z API {vrep['filled_from_api']}, "
        f"zostalo {vrep['empty_after']['fields']}"
    )
    if vrep.get("remaining"):
        _safe_print("Nadal puste (brak w JSON i na stronach):")
        for g in vrep["remaining"]:
            _safe_print(
                f"  {g['date']} {g['home']} - {g['away']}: "
                f"{len(g['missing'] or [])} pol"
            )
    leftover = int((vrep.get("empty_after") or {}).get("fields") or 0)
    if require_complete and leftover > 0:
        raise SystemExit(
            f"Po weryfikacji zostalo {leftover} pustych pol statystyk — Excel nie jest kompletny."
        )
    changed = (vrep["filled_from_json"] + vrep["filled_from_api"]) > 0
    return mecze_x, preds_x, changed


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Predykcje 2026")
    parser.add_argument(
        "--od",
        default=FROM_DATE.strftime("%d/%m/%Y"),
        type=parse_od_date,
        help="Poczatek zakresu meczow w Excelu (dd/mm/yyyy), domyslnie 13/08/2026",
    )
    parser.add_argument(
        "--no-upcoming",
        action="store_true",
        help="Nie dociagaj nadchodzacych meczow z FootyStats",
    )
    parser.add_argument("--upcoming-days", type=int, default=7, help="Ile dni do przodu (BBC/FootyStats)")
    parser.add_argument("--refresh-upcoming", action="store_true", help="Pomin cache FootyStats")
    parser.add_argument(
        "--fill-missing",
        action="store_true",
        help="Zawsze wlaczone: JSON + Serper/Claude + weryfikacja Excela (flaga zostaje dla CI)",
    )
    parser.add_argument(
        "--send-mail",
        action="store_true",
        help="Wyslij predykcje_2026.xlsx na MAIL_TO (Gmail SMTP)",
    )
    args = parser.parse_args(argv)

    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

    df_year, df_history = load_2026_data()
    n_year = len(df_year)
    if not args.no_upcoming:
        try:
            import upcoming as up

            fixtures = up.fetch_upcoming_fixtures(
                days=args.upcoming_days,
                refresh=args.refresh_upcoming,
                history=df_history,
            )
            if not fixtures.empty:
                fixtures[COL_DATA] = pd.to_datetime(fixtures[COL_DATA], dayfirst=True, errors="coerce")
                fixtures = _to_numeric(fixtures)
                fixtures = _attach_goals(fixtures)
                before = len(df_year)
                n_scored = int(
                    fixtures[COL_RESULT].astype(str).str.match(r"^\d+:\d+$").sum()
                ) if COL_RESULT in fixtures.columns else 0
                df_year = up.merge_upcoming(df_year, fixtures)
                _safe_print(
                    f"BBC: {len(fixtures)} meczow "
                    f"(FT {n_scored}, dodano {len(df_year) - before})"
                )
            else:
                _safe_print("BBC: 0 meczow lig Aleksa")
        except Exception as exc:
            logger.warning("Upcoming pominiete: %s", exc)
            _safe_print(f"Upcoming pominiete: {exc}")
    df_2026 = filter_from_date(df_year, args.od)
    validate_from_date(df_2026, args.od, label="Mecze")
    with_score = 0
    if COL_RESULT in df_history.columns:
        with_score = int(df_history[COL_RESULT].astype(str).str.match(r"^\d+:\d+$").sum())
    _safe_print(
        f"Mecze od {args.od.strftime('%d/%m/%Y')}: {len(df_2026)} "
        f"(z {n_year} w 2026) | historii: {len(df_history)} | z wynikiem: {with_score}"
    )
    if df_2026.empty:
        raise SystemExit(
            f"Brak meczow od {args.od.strftime('%d/%m/%Y')} — nie mozna wygenerowac predykcji."
        )

    df_2026, df_history, df_year = _fill_played_from_json_and_api(
        df_2026,
        df_history,
        df_year,
        from_date=args.od,
        require_complete=True,
    )

    team_avg = compute_team_averages(df_history)
    league_avg = compute_league_averages(df_2026)
    _safe_print("Buduje predykcje (forma + O/U)...")
    predictions = build_predictions(df_2026, df_history, use_as_of=True)
    validate_from_date(predictions, args.od, label="Predykcje")

    _safe_print("Backtest 2025...")
    bt_detail, bt_summary = run_backtest(df_history, year=2025)
    if not bt_summary.empty:
        for _, r in bt_summary.head(8).iterrows():
            _safe_print(f"  {r['metryka']}: {r['wartosc']}")

    path = export_excel(
        df_2026,
        predictions,
        team_avg,
        league_avg,
        backtest_detail=bt_detail,
        backtest_summary=bt_summary,
        from_date=args.od,
    )

    mecze_x, preds_x, changed = _verify_exported_stats(
        path,
        from_date=args.od,
        require_complete=True,
    )
    if changed and mecze_x is not None:
        df_2026 = mecze_x
        if preds_x is not None:
            predictions = preds_x
        path = export_excel(
            df_2026,
            predictions,
            team_avg,
            league_avg,
            backtest_detail=bt_detail,
            backtest_summary=bt_summary,
            from_date=args.od,
            path=path,
        )
        _safe_print(f"Excel po weryfikacji: {path}")

    _safe_print(f"Predykcje: {len(predictions)} meczow")
    if "status" in predictions.columns:
        _safe_print(f"Status: {predictions['status'].value_counts().to_dict()}")
    if "metoda" in predictions.columns:
        _safe_print(f"Metody: {predictions['metoda'].value_counts().to_dict()}")
    _safe_print(f"Plik: {path}")
    if args.send_mail:
        try:
            import send_mail as mail

            info = mail.send_excel(path)
            _safe_print(f"Mail: {Path(info['file']).name} → {info['to']}")
        except Exception as exc:
            logger.warning("Wysylka maila pominieta: %s", exc)
            _safe_print(f"Wysylka maila pominieta: {exc}")


if __name__ == "__main__":
    main()
