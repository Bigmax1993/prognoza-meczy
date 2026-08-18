# -*- coding: utf-8 -*-
"""Scraper meczów z footystats.org → cache JSON + Excel."""
from __future__ import annotations

import argparse
import json
import logging
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse

import pandas as pd
import requests
from bs4 import BeautifulSoup, NavigableString
from openpyxl import Workbook

# ---------------------------------------------------------------------------
# Ścieżki / stałe
# ---------------------------------------------------------------------------

OUTPUT_DIR = Path(__file__).resolve().parent
CACHE_DIR = OUTPUT_DIR / "cache"
LOGS_DIR = OUTPUT_DIR / "logs"
MATCHES_CACHE = CACHE_DIR / "matches_cache.json"
COOKIES_CACHE = CACHE_DIR / "cookies_cache.json"
HTML_CACHE = CACHE_DIR / "html_cache.json"

URL = "https://footystats.org/ru/"
BASE_ORIGIN = "https://footystats.org"

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)

logger = logging.getLogger("footystats_scraper")


# ---------------------------------------------------------------------------
# Cloudflare
# ---------------------------------------------------------------------------

_CF_TITLE_MARKERS = (
    "just a moment",
    "один момент",
    "attention required",
)
_CF_HTML_MARKERS = (
    "cf-turnstile",
    "challenge-platform",
    "challenges.cloudflare.com",
    "checking your browser",
    "выполнение проверки безопасности",
    "cdn-cgi/challenge",
    "cf-browser-verification",
)


def _is_cloudflare_block(html: str, title: str = "") -> bool:
    """Wykrywa stronę wyzwania Cloudflare (EN/RU)."""
    title_l = (title or "").lower()
    html_l = (html or "").lower()
    if any(m in title_l for m in _CF_TITLE_MARKERS):
        return True
    if "cloudflare" in title_l and "attention" in title_l:
        return True
    if any(m in html_l for m in _CF_HTML_MARKERS):
        return True
    return False


def _extract_title(html: str) -> str:
    try:
        soup = BeautifulSoup(html, "lxml")
        if soup.title and soup.title.string:
            return soup.title.string.strip()
    except Exception:
        pass
    m = re.search(r"<title[^>]*>(.*?)</title>", html or "", re.I | re.S)
    return m.group(1).strip() if m else ""


# ---------------------------------------------------------------------------
# Cache JSON
# ---------------------------------------------------------------------------

def _ensure_dirs() -> None:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def load_json_cache(path: Path) -> dict | None:
    """Wczytuje JSON; None gdy brak pliku lub uszkodzony."""
    try:
        path = Path(path)
        if not path.exists():
            return None
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def _write_json(path: Path, payload: dict) -> Path:
    _ensure_dirs()
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return path


def _cache_age_ok(cached_at: str, ttl_minutes: int) -> bool:
    try:
        ts = datetime.fromisoformat(cached_at)
    except Exception:
        return False
    return datetime.now() - ts <= timedelta(minutes=ttl_minutes)


def save_matches_cache(match_date: str, rows: list[dict], source: str) -> Path:
    payload = {
        "cached_at": datetime.now().isoformat(timespec="seconds"),
        "url": URL,
        "match_date": match_date,
        "source": source,
        "count": len(rows),
        "matches": rows,
    }
    path = _write_json(MATCHES_CACHE, payload)
    logger.info(
        "Zapisano cache meczów (%s) — %s pozycji → %s",
        source,
        len(rows),
        path.name,
    )
    return path


def load_matches_cache(ttl_minutes: int = 30) -> tuple[str, list[dict]] | None:
    data = load_json_cache(MATCHES_CACHE)
    if not data:
        return None
    if not _cache_age_ok(data.get("cached_at", ""), ttl_minutes):
        logger.info("Cache meczów wygasł (TTL=%s min)", ttl_minutes)
        return None
    match_date = data.get("match_date") or ""
    rows = data.get("matches") or []
    try:
        age = datetime.now() - datetime.fromisoformat(data["cached_at"])
        logger.info(
            "Cache JSON — %s meczów (wiek: %s, plik: %s)",
            len(rows),
            age,
            Path(MATCHES_CACHE).name,
        )
    except Exception:
        pass
    return match_date, rows


def save_cookies_cache(cookies: list[dict]) -> Path:
    payload = {
        "cached_at": datetime.now().isoformat(timespec="seconds"),
        "cookies": cookies,
    }
    return _write_json(COOKIES_CACHE, payload)


def load_cookies_cache(ttl_minutes: int = 60) -> list | None:
    data = load_json_cache(COOKIES_CACHE)
    if not data:
        return None
    if not _cache_age_ok(data.get("cached_at", ""), ttl_minutes):
        return None
    return data.get("cookies")


def save_html_cache(html: str, source: str) -> Path:
    payload = {
        "cached_at": datetime.now().isoformat(timespec="seconds"),
        "url": URL,
        "source": source,
        "length": len(html or ""),
        "html": html or "",
    }
    return _write_json(HTML_CACHE, payload)


# ---------------------------------------------------------------------------
# Cookies → Playwright
# ---------------------------------------------------------------------------

def _cookies_for_playwright(session: Any, url: str) -> list[dict]:
    """Konwertuje cookies sesji (dict-like lub jar) na listę dla Playwright."""
    host = urlparse(url).hostname or "footystats.org"
    result: list[dict] = []
    cookies_obj = getattr(session, "cookies", None)
    if cookies_obj is None:
        return result

    jar = getattr(cookies_obj, "jar", None)
    if jar is not None:
        for c in jar:
            result.append(
                {
                    "name": getattr(c, "name", ""),
                    "value": getattr(c, "value", ""),
                    "domain": getattr(c, "domain", None) or host,
                    "path": getattr(c, "path", None) or "/",
                }
            )
        return result

    # dict-like (np. RequestsCookieJar lub zwykły dict)
    try:
        items = list(cookies_obj.items())
    except Exception:
        try:
            items = [(c.name, c.value) for c in cookies_obj]
        except Exception:
            items = []

    for name, value in items:
        result.append(
            {
                "name": str(name),
                "value": str(value),
                "domain": host,
                "path": "/",
            }
        )
    return result


# ---------------------------------------------------------------------------
# Fetch: requests / playwright / curl_cffi
# ---------------------------------------------------------------------------

def fetch_with_requests(url: str) -> str | None:
    """Prosty GET; None przy błędzie, nie-200 lub Cloudflare."""
    try:
        resp = requests.get(
            url,
            timeout=45,
            headers={"User-Agent": USER_AGENT, "Accept-Language": "ru,en;q=0.8"},
        )
        if resp.status_code != 200:
            logger.info("requests: status=%s", resp.status_code)
            return None
        title = _extract_title(resp.text)
        if _is_cloudflare_block(resp.text, title):
            logger.info("requests: wykryto Cloudflare (%s)", title)
            return None
        return resp.text
    except Exception as exc:
        logger.info("requests: wyjątek — %s", exc)
        return None


def _playwright_launch_args() -> list[str]:
    """Flagi Chromium pod cichy tryb w tle (bez okna UI)."""
    return [
        "--disable-gpu",
        "--no-sandbox",
        "--disable-dev-shm-usage",
        "--disable-extensions",
        "--disable-background-networking",
        "--mute-audio",
        "--window-position=-2400,-2400",
        "--window-size=1280,720",
    ]


# Selektory / teksty przyciskow zgody (cookies, privacy, age gate)
_COOKIE_BUTTON_SELECTORS = (
    "#onetrust-accept-btn-handler",
    "#accept-recommended-btn-handler",
    "button#onetrust-accept-btn-handler",
    ".onetrust-accept-btn-handler",
    "#CybotCookiebotDialogBodyLevelButtonLevelOptinAllowAll",
    "#CybotCookiebotDialogBodyButtonAccept",
    "button[data-testid='cookie-policy-dialog-accept-button']",
    "button[aria-label*='Accept' i]",
    "button[aria-label*='Akceptuj' i]",
    "button[aria-label*='Принять' i]",
    "button[aria-label*='Agree' i]",
    "#cookie-accept",
    "#acceptCookies",
    ".cookie-accept",
    ".accept-cookies",
    ".cc-btn.cc-dismiss",
    ".cc-allow",
    ".cc-accept",
    ".js-cookie-accept",
    "button.cookie-consent-accept",
    "[data-action='accept']",
    "[data-cookie-accept]",
)

_COOKIE_BUTTON_TEXTS = (
    "Accept all",
    "Accept All",
    "Accept cookies",
    "Accept Cookies",
    "I agree",
    "I Agree",
    "Allow all",
    "Allow All",
    "Got it",
    "OK",
    "Agree",
    "Akceptuj",
    "Akceptuję",
    "Akceptuj wszystkie",
    "Zgadzam się",
    "Zaakceptuj",
    "Принять",
    "Принять все",
    "Согласен",
    "Соглашаюсь",
    "Хорошо",
    "Принимаю",
    "Принять cookie",
    "Принять Cookie",
)


def _accept_cookies_and_consents(page) -> bool:
    """
    Klika 'Accept cookies' / 'Принять' / podobne banery zgody.
    Zwraca True, jesli udalo sie kliknac przynajmniej jeden przycisk.
    """
    clicked = False

    # 1) Znane selektory CMP / cookie bannerow
    for sel in _COOKIE_BUTTON_SELECTORS:
        try:
            loc = page.locator(sel).first
            if loc.count() == 0:
                continue
            if not loc.is_visible(timeout=800):
                continue
            loc.click(timeout=2000)
            logger.info("Playwright: kliknieto zgode cookies (selector: %s)", sel)
            clicked = True
            page.wait_for_timeout(500)
            break
        except Exception:
            continue

    # 2) Przyciski / linki po widocznym tekscie (EN / PL / RU)
    if not clicked:
        for text in _COOKIE_BUTTON_TEXTS:
            try:
                loc = page.get_by_role("button", name=re.compile(rf"^\s*{re.escape(text)}\s*$", re.I))
                if loc.count() == 0:
                    loc = page.get_by_text(re.compile(rf"^\s*{re.escape(text)}\s*$", re.I))
                if loc.count() == 0:
                    continue
                btn = loc.first
                if not btn.is_visible(timeout=800):
                    continue
                btn.click(timeout=2000)
                logger.info("Playwright: kliknieto zgode cookies (tekst: %s)", text)
                clicked = True
                page.wait_for_timeout(500)
                break
            except Exception:
                continue

    # 3) iframe z cookie bannerem (Cookiebot / OneTrust)
    if not clicked:
        try:
            for frame in page.frames:
                if frame == page.main_frame:
                    continue
                for sel in _COOKIE_BUTTON_SELECTORS[:8]:
                    try:
                        loc = frame.locator(sel).first
                        if loc.count() and loc.is_visible(timeout=500):
                            loc.click(timeout=2000)
                            logger.info("Playwright: kliknieto cookies w iframe (%s)", sel)
                            clicked = True
                            break
                    except Exception:
                        continue
                if clicked:
                    break
        except Exception:
            pass

    # 4) Zamknij ewentualny overlay / modal bliski cookies
    if clicked:
        try:
            page.keyboard.press("Escape")
            page.wait_for_timeout(300)
        except Exception:
            pass

    return clicked


def fetch_with_playwright(url: str) -> str | None:
    """
    Pobiera stronę przez Chromium (Playwright) w tle.
    Klika akceptację cookies; najpierw headless, potem off-screen vs Cloudflare.
    """
    try:
        from playwright.sync_api import sync_playwright
    except Exception as exc:
        logger.info("playwright niedostepny: %s", exc)
        return None

    def _stealth(page) -> None:
        page.add_init_script(
            """
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            Object.defineProperty(navigator, 'languages', { get: () => ['ru-RU', 'ru', 'en-US', 'en'] });
            Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] });
            window.chrome = { runtime: {} };
            """
        )

    def _load(headless: bool) -> tuple[str, str] | None:
        context = None
        try:
            profile_dir = OUTPUT_DIR / (".pw_profile" if headless else ".pw_profile_os")
            profile_dir.mkdir(parents=True, exist_ok=True)
            cached_cookies = load_cookies_cache(ttl_minutes=120) or []
            args = _playwright_launch_args()
            if not headless:
                args = [
                    a
                    for a in args
                    if not a.startswith("--window-position") and a != "--disable-gpu"
                ] + ["--window-position=-32000,-32000", "--start-minimized"]

            with sync_playwright() as p:
                launch_kwargs = dict(
                    user_data_dir=str(profile_dir),
                    headless=headless,
                    args=args,
                    user_agent=USER_AGENT,
                    locale="ru-RU",
                    viewport={"width": 1365, "height": 900},
                    java_script_enabled=True,
                    ignore_https_errors=True,
                )
                try:
                    context = p.chromium.launch_persistent_context(
                        channel="chrome", **launch_kwargs
                    )
                except Exception:
                    context = p.chromium.launch_persistent_context(**launch_kwargs)

                if cached_cookies:
                    try:
                        context.add_cookies(cached_cookies)
                    except Exception:
                        pass

                page = context.pages[0] if context.pages else context.new_page()
                _stealth(page)
                page.goto(url, wait_until="domcontentloaded", timeout=120_000)

                # od razu sprobuj zaakceptowac cookies
                try:
                    page.wait_for_timeout(1200)
                    _accept_cookies_and_consents(page)
                except Exception as exc:
                    logger.debug("cookies click (pierwszy): %s", exc)

                deadline_ms = 75_000 if not headless else 45_000
                stepped = 0
                while stepped < deadline_ms:
                    title = page.title()
                    html = page.content()
                    # ponawiaj klik cookies w trakcie ladowania (banner moze dojsc pozniej)
                    try:
                        _accept_cookies_and_consents(page)
                    except Exception:
                        pass
                    if not _is_cloudflare_block(html, title):
                        if page.query_selector("div.league a.match, a.match, div.league"):
                            break
                        if "footystats" in (title or "").lower():
                            break
                    page.wait_for_timeout(2500)
                    stepped += 2500

                try:
                    page.wait_for_selector("div.league a.match, a.match", timeout=20_000)
                except Exception:
                    pass

                # ostatnia proba cookies przed zrzutem HTML
                try:
                    _accept_cookies_and_consents(page)
                    page.wait_for_timeout(400)
                except Exception:
                    pass

                html = page.content()
                title = page.title()
                try:
                    pw_cookies = context.cookies()
                    if pw_cookies:
                        save_cookies_cache(
                            [
                                {
                                    "name": c.get("name", ""),
                                    "value": c.get("value", ""),
                                    "domain": c.get("domain", "footystats.org"),
                                    "path": c.get("path", "/"),
                                }
                                for c in pw_cookies
                            ]
                        )
                except Exception:
                    pass
                context.close()
                context = None

            if _is_cloudflare_block(html, title):
                return None
            return html, title
        except Exception as exc:
            logger.info("playwright(%s): %s", "headless" if headless else "offscreen", exc)
            return None
        finally:
            if context is not None:
                try:
                    context.close()
                except Exception:
                    pass

    got = _load(headless=True)
    if got:
        html, _title = got
        logger.info("playwright: headless OK (%s bajtow)", len(html))
        return html

    logger.info("playwright: headless zablokowany — proba off-screen")
    got = _load(headless=False)
    if got:
        html, _title = got
        logger.info("playwright: off-screen OK (%s bajtow)", len(html))
        return html

    logger.info("playwright: Cloudflare nadal aktywny")
    return None


def _curl_cffi_session(url: str, impersonate: str = "chrome124"):
    """Tworzy sesję curl_cffi i zwraca (session, html) albo rzuca wyjątek."""
    from curl_cffi import requests as curl_requests

    session = curl_requests.Session(impersonate=impersonate)
    resp = session.get(
        url,
        timeout=60,
        headers={"Accept-Language": "ru,en;q=0.8"},
    )
    if getattr(resp, "status_code", 0) != 200:
        raise RuntimeError(f"curl_cffi status={getattr(resp, 'status_code', '?')}")
    html = getattr(resp, "text", "") or ""
    title = _extract_title(html)
    if _is_cloudflare_block(html, title):
        raise RuntimeError(f"curl_cffi Cloudflare: {title}")
    return session, html


def fetch_with_curl_cffi(url: str) -> str | None:
    """Fetch przez curl_cffi (impersonacja Chrome); zapisuje cookies."""
    try:
        session, html = _curl_cffi_session(url)
        cookies = _cookies_for_playwright(session, url)
        if cookies:
            save_cookies_cache(cookies)
        return html
    except Exception as exc:
        logger.info("curl_cffi: wyjątek — %s", exc)
        return None


def fetch_html(url: str) -> tuple[str, str]:
    """
    Pobiera HTML kolejno: playwright (headless) → curl_cffi → requests.
    Zapisuje html_cache. SystemExit(1) gdy wszystkie metody zawiodą.
    """
    strategies = (
        ("playwright", fetch_with_playwright),
        ("curl_cffi", fetch_with_curl_cffi),
        ("requests", fetch_with_requests),
    )
    for source, fn in strategies:
        logger.info("Próba pobrania: %s", source)
        html = fn(url)
        if html:
            save_html_cache(html, source)
            logger.info("Pobrano HTML przez %s (%s bajtów)", source, len(html))
            return html, source

    logger.error("Wszystkie metody pobierania zawiodły")
    raise SystemExit(1)


# ---------------------------------------------------------------------------
# Parsowanie HTML
# ---------------------------------------------------------------------------

def _text(el) -> str:
    if el is None:
        return ""
    return el.get_text(" ", strip=True)


def _odd_value(span) -> str:
    """Pierwsza liczba/tekst z spanu odds (bez hover-modal-content)."""
    if span is None:
        return ""
    parts: list[str] = []
    for child in span.children:
        if isinstance(child, NavigableString):
            t = str(child).strip()
            if t:
                parts.append(t)
    if parts:
        return parts[0].split()[0]
    raw = span.get_text(" ", strip=True)
    if not raw:
        return ""
    return raw.split()[0]


def _strip_country_prefix(league_title: str, country: str = "") -> str:
    title = (league_title or "").strip()
    if " - " in title:
        return title.split(" - ", 1)[1].strip()
    if country and title.startswith(country):
        return title[len(country) :].lstrip(" -–—").strip()
    return title


def parse_matches(html: str) -> tuple[str, list[dict]]:
    """Parsuje listę meczów ze struktury FootyStats (league / match row)."""
    soup = BeautifulSoup(html or "", "lxml")

    match_date = datetime.now().strftime("%d/%m")
    title_el = soup.select_one("h2.postTitle")
    if title_el:
        m = re.search(r"(\d{1,2}/\d{1,2})", title_el.get_text(" ", strip=True))
        if m:
            match_date = m.group(1)

    rows: list[dict] = []
    for league in soup.select("div.league"):
        country_raw = _text(league.select_one(".league-country"))
        country = country_raw.rstrip("-–— ").strip()
        league_title = _text(league.select_one(".league-title"))
        liga = _strip_country_prefix(league_title, country)

        for match in league.select("a.match"):
            href = match.get("href") or ""
            link = urljoin(BASE_ORIGIN, href) if href else ""

            time_el = match.select_one(".timezone-convert-match-regular")
            godzina = _text(time_el)

            status_el = match.select_one(".match-time-soon")
            status = ""
            if status_el is not None:
                status = status_el.get("data-match-status") or ""

            home = match.select_one(".team.home")
            away = match.select_one(".team.away")
            gospodarz = _text(
                home.select_one(".hover-modal-ajax-team") if home else None
            )
            forma_gosp = _text(home.select_one(".form-box") if home else None)
            gosc = _text(
                away.select_one(".hover-modal-ajax-team") if away else None
            )
            forma_gosc = _text(away.select_one(".form-box") if away else None)

            odd_spans = match.select(".stat.odds .hover-modal-parent")
            kurs1 = _odd_value(odd_spans[0]) if len(odd_spans) > 0 else ""
            kurs_x = _odd_value(odd_spans[1]) if len(odd_spans) > 1 else ""
            kurs2 = _odd_value(odd_spans[2]) if len(odd_spans) > 2 else ""

            rows.append(
                {
                    "Data": match_date,
                    "Kraj": country,
                    "Liga": liga,
                    "Godzina": godzina,
                    "Gospodarz": gospodarz,
                    "Forma gospodarza": forma_gosp,
                    "Gość": gosc,
                    "Forma gościa": forma_gosc,
                    "Kurs 1": kurs1,
                    "Kurs X": kurs_x,
                    "Kurs 2": kurs2,
                    "Status": status,
                    "Link": link,
                }
            )

    return match_date, rows


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def save_excel(rows: list[dict], path: Path | str) -> Path:
    """Zapisuje arkusz 'Mecze'; puste → A1='Brak meczów'."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Mecze"

    if not rows:
        ws["A1"] = "Brak meczów"
        wb.save(path)
        return path

    headers = list(rows[0].keys())
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col, header)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, key in enumerate(headers, start=1):
            ws.cell(r_idx, c_idx, row.get(key, ""))
    wb.save(path)
    logger.info("Zapisano Excel: %s", path)
    return path


# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

def setup_logging(level: str | int = "INFO") -> Path:
    """Konfiguruje logger do pliku scraper_YYYY-MM-DD.log."""
    _ensure_dirs()
    log_path = LOGS_DIR / f"scraper_{datetime.now():%Y-%m-%d}.log"

    if isinstance(level, str):
        level_no = getattr(logging, level.upper(), logging.INFO)
    else:
        level_no = int(level)

    logger.handlers.clear()
    logger.setLevel(level_no)
    logger.propagate = False

    fmt = logging.Formatter(
        fmt="%(asctime)s | %(levelname)-8s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setLevel(level_no)
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    sh = logging.StreamHandler(sys.stdout)
    sh.setLevel(level_no)
    sh.setFormatter(fmt)
    logger.addHandler(sh)

    return log_path


# ---------------------------------------------------------------------------
# Pandas preview (delegacja do preview_pandas)
# ---------------------------------------------------------------------------

def preview_dataframe(rows: list[dict], n: int = 10) -> pd.DataFrame:
    """Buduje DataFrame, wypisuje head(n) i zapisuje preview_head.*."""
    from preview_pandas import preview_dataframe as _preview

    _ensure_dirs()
    logger.info("Podgląd DataFrame — shape=%s, head(%s)", len(rows), n)
    df = _preview(rows, n=n, output_dir=OUTPUT_DIR, prefix="preview_head", save=True, show=True)
    logger.info("Podgląd zapisany: preview_head.txt | preview_head.csv")
    return df


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Scraper FootyStats → Excel")
    parser.add_argument(
        "--refresh",
        action="store_true",
        help="Pomiń cache i pobierz stronę na nowo",
    )
    parser.add_argument(
        "--log-level",
        default="INFO",
        help="Poziom logowania (DEBUG/INFO/WARNING/...)",
    )
    args = parser.parse_args(argv)

    log_path = setup_logging(args.log_level)
    logger.info("Start scrapera | logi: %s", log_path)

    try:
        rows: list[dict]
        match_date: str
        source: str

        if not args.refresh:
            cached = load_matches_cache(ttl_minutes=30)
            if cached is not None:
                match_date, rows = cached
                source = "cache"
            else:
                cached = None
        else:
            cached = None

        if args.refresh or cached is None:
            html, source = fetch_html(URL)
            match_date, rows = parse_matches(html)
            save_matches_cache(match_date, rows, source)

        logger.info(
            "Znaleziono %s meczów (data: %s, źródło: %s)",
            len(rows),
            match_date,
            source,
        )

        excel_name = f"footystats_mecze_{match_date.replace('/', '-')}.xlsx"
        excel_path = OUTPUT_DIR / excel_name
        save_excel(rows, excel_path)
        preview_dataframe(rows, n=10)
    except SystemExit:
        raise
    except Exception:
        logger.exception("Nieoczekiwany błąd scrapera")
        raise


if __name__ == "__main__":
    main()
