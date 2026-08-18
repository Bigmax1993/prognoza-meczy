# -*- coding: utf-8 -*-
"""Predykcje MAX — rozszerzony system predykcji (pandas + Poisson-lite).

Warstwy:
  1. Heurystyka pandas + forma 3/5/10 + H2H + ranking (bez leakage)
  2. Kalibracja progów O/U per liga (train 2024 → valid 2025)
  3. Opcjonalnie ML (--ml) — sklearn LogisticRegression

Uruchomienie:
  python predykcje_max.py
  python predykcje_max.py --ml
"""
from __future__ import annotations

import argparse
import logging
import math
import re
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

import predykcje as base
from monthly_summary import compute_monthly_team_summary
from preview_pandas import ALEKS_XLSX, resolve_aleks_path

ROOT = Path(__file__).resolve().parent
OUT_XLSX = ROOT / "predykcje_max_2026.xlsx"
AUDYT_MD = ROOT / "AUDYT_DANYCH.md"

logger = logging.getLogger("predykcje_max")

COL_LIGA = base.COL_LIGA
COL_DATA = base.COL_DATA
COL_HOME = base.COL_HOME
COL_AWAY = base.COL_AWAY
COL_BTTS = base.COL_BTTS
COL_RESULT = base.COL_RESULT
STAT_COLS = base.STAT_COLS

MIN_MATCHES = 5
FORM_WINDOWS = (3, 5, 10)
GOAL_LINES = (1.5, 2.5, 3.5)
CORNER_LINES = (8.5, 9.5, 10.5)
CARD_LINES = (2.5, 3.5, 4.5)
TOP_N_TYPOW = 20

NORDIC_LEAGUES = {"Allsvenskan", "Eliteserien", "Super League"}

# --- Ukrainizacja eksportu Excel ---
UA_COL: dict[str, str] = {
    "druzyna": "команда",
    "miesiac": "місяць",
    "mecze": "матчі",
    "status": "статус",
    "powod": "причина",
    "przewidywany_zwyciezca": "прогноз_переможець",
    "prawdopodobienstwo_gospodarz": "ймовірність_господар",
    "prawdopodobienstwo_remis": "ймовірність_нічия",
    "prawdopodobienstwo_gosc": "ймовірність_гість",
    "przewidywany_wynik": "прогноз_рахунок",
    "przewidywany_wynik_poisson": "прогноз_рахунок_poisson",
    "exp_gole_gospodarz": "очікувані_голи_господар",
    "exp_gole_gosc": "очікувані_голи_гість",
    "przewidywane_btts": "прогноз голів",
    "przewidywane_btts_pct": "прогноз_оз_відсоток",
    "exp_rozne": "очікувані_кутові",
    "exp_zolte": "очікувані_жовті",
    "exp_faule": "очікувані_фоли",
    "linia_faule": "лінія_фоли",
    "linia_rozne": "лінія_кутові",
    "linia_zolte": "лінія_жовті",
    "forma_gospodarz_mecze": "форма_господар_матчі",
    "forma_gosc_mecze": "форма_гість_матчі",
    "metoda": "метод",
    "pewnosc": "впевненість",
    "pewnosc_score": "бал_впевненості",
    "min_meczy_historii": "мін_матчів_історії",
    "ma_statystyki": "є_статистика",
    "predykcja_rozne": "прогноз_кутові",
    "predykcja_zolte": "прогноз_жовті",
    "predykcja_faule": "прогноз_фоли",
    "rozne_gospodarz_exp": "кутові_господар_очікувані",
    "rozne_gosc_exp": "кутові_гість_очікувані",
    "pred_rozne_gospodarz_4_5": "прогноз_кутові_господар_4_5",
    "pred_rozne_gosc_4_5": "прогноз_кутові_гість_4_5",
    "handicap_0": "гандикап_0",
    "handicap_h_minus_0_5": "гандикап_господар_мінус_0_5",
    "value_flag": "ознака_цінності",
    "actual_zwyciezca": "факт_переможець",
    "actual_btts": "факт_оз",
    "actual_rozne": "факт_кутові",
    "actual_zolte": "факт_жовті",
    "actual_gole": "факт_голи",
    "trafiony_1x2": "влучено_1x2",
    "metryka": "метрика",
    "wartosc": "значення",
    "baseline_predykcje": "базова_модель",
    "predykcje_max": "модель_макс",
    "delta_pp": "різниця_п_п",
    "lepszy": "краща",
    "btts_yes_pct": "оз_так_відсоток",
    "mecze_z_wynikiem": "матчі_з_рахунком",
    "srednia_goli_strzelonych": "середн_забиті",
    "srednia_goli_straconych": "середн_пропущені",
    "srednia_faule": "середн_фоли",
    "srednia_rozne": "середн_кутові",
    "srednia_zolte_kartki": "середн_жовті",
    "srednia_strzaly": "середн_удари",
    "srednia_strzaly_w_stwor": "середн_удари_в_площину",
    "as_of": "станом_на",
    "ranking_h": "рейтинг_господар",
    "ranking_a": "рейтинг_гість",
    "punkty_h": "очки_господар",
    "punkty_a": "очки_гість",
    "days_since_h": "днів_від_матчу_господар",
    "days_since_a": "днів_від_матчу_гість",
    "home_advantage_ligi": "перевага_господарів_ліги",
    "h2h_mecze": "очні_матчі",
    "h2h_home_wins": "очні_перемоги_господар",
    "h2h_avg_goli": "очні_середн_голи",
    "h2h_btts_pct": "очні_оз_відсоток",
    "info": "інформація",
}

UA_VAL: dict[str, str] = {
    "over": "більше",
    "under": "менше",
    "remis": "нічия",
    "wysoka": "висока",
    "srednia": "середня",
    "niska": "низька",
    "ok": "активний",
    "pomin": "пропущено",
    "za_malo_danych": "замало_даних",
    "nordic_bez_stat": "без_статистики",
    "model_wynikow": "модель_голів",
    "forma_5": "форма_5",
    "heurystyka_statystyk": "евристика_стат",
    "max": "макс",
    "baseline": "базова",
    "max_only": "тільки_макс",
    "predykcje_max": "прогнози_макс",
    "brak_danych": "немає_даних",
    "hit_rate_1x2_pct": "точність_1x2_відсоток",
    "hit_rate_btts_pct": "точність_оз_відсоток",
    "hit_rate_gole_2_5_pct": "точність_голи_2_5_відсоток",
    "hit_rate_rozne_9_5_pct": "точність_кутові_9_5_відсоток",
    "hit_rate_zolte_3_5_pct": "точність_жовті_3_5_відсоток",
    "mecze_razem": "матчів_усього",
    "mecze_ocenione": "матчів_оцінено",
    "pominiete": "пропущено",
    "rok": "рік",
    "model": "модель",
}

UA_SHEET_SUMMARY = {
    "team_avg": "Середні_команди",
    "league_avg": "Середні_ліги",
    "monthly": "Підсумок_місяця",
    "backtest": "Бектест_2025",
    "backtest_detail": "Бектест_деталі",
    "benchmark": "Порівняння",
    "audyt": "Аудит",
    "top_typy": "Топ_типи",
    "charts": "Графіки",
}


def _ua_column(name: str) -> str:
    if name in UA_COL:
        return UA_COL[name]
    if name.startswith("predykcja_gole_"):
        line = name.replace("predykcja_gole_", "").replace(".", "_")
        return f"прогноз_голи_{line}"
    if name.startswith("ou_gole_") and name.endswith("_pct"):
        line = name.replace("ou_gole_", "").replace("_pct", "").replace(".", "_")
        return f"голи_{line}_більше_відсоток"
    if name.startswith("predykcja_rozne_"):
        line = name.replace("predykcja_rozne_", "").replace(".", "_")
        return f"прогноз_кутові_{line}"
    if name.startswith("predykcja_zolte_"):
        line = name.replace("predykcja_zolte_", "").replace(".", "_")
        return f"прогноз_жовті_{line}"
    if name.startswith("forma_") and name.endswith("_gf"):
        parts = name.split("_")
        return f"форма_{parts[1]}_{parts[2]}_забиті"
    if name.startswith("forma_") and name.endswith("_mecze"):
        parts = name.split("_")
        side = "г" if parts[2] == "h" else "гість"
        return f"форма_{parts[1]}_{side}_матчі"
    if name.startswith("srednia_"):
        return "середн_" + name.replace("srednia_", "")
    return name


def _ua_cell(value: object) -> object:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return value
    text = str(value).strip()
    return UA_VAL.get(text, UA_VAL.get(text.casefold(), value))


def drop_empty_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Usuwa kolumny w 100% puste (NaN, pusty string, 'nan')."""
    if df.empty:
        return df

    def _has_value(val: object) -> bool:
        if val is None:
            return False
        if isinstance(val, float) and pd.isna(val):
            return False
        text = str(val).strip().casefold()
        return text not in {"", "nan", "none", "<na>", "nat"}

    keep: list[Any] = []
    for col in df.columns:
        if any(_has_value(v) for v in df[col].tolist()):
            keep.append(col)
    return df[keep].copy() if keep else df.iloc[:, 0:0].copy()


def translate_df_ua(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    out = df.copy()
    out.columns = [_ua_column(str(c)) for c in out.columns]
    for idx in range(out.shape[1]):
        series = out.iloc[:, idx]
        if not pd.api.types.is_numeric_dtype(series):
            out.iloc[:, idx] = series.map(_ua_cell)
    return out


def _table_display_name(base: str, used: set[str]) -> str:
    clean = re.sub(r"[^a-zA-Z0-9_]", "_", base)
    if not clean or clean[0].isdigit():
        clean = "T_" + (clean or "data")
    name = clean[:25]
    i = 1
    while name in used:
        suffix = f"_{i}"
        name = clean[: 25 - len(suffix)] + suffix
        i += 1
    used.add(name)
    return name



def _safe_print(text: str) -> None:
    base._safe_print(text)


def has_league_stats(league: str, df: pd.DataFrame) -> bool:
    if league in NORDIC_LEAGUES:
        return False
    sub = df[df[COL_LIGA] == league]
    if sub.empty or "кутові" not in sub.columns:
        return False
    return pd.to_numeric(sub["кутові"], errors="coerce").notna().any()


def _mean(vals: list[float]) -> float | None:
    return sum(vals) / len(vals) if vals else None


def poisson_pmf(k: int, lam: float) -> float:
    if lam <= 0:
        return 1.0 if k == 0 else 0.0
    return math.exp(-lam) * lam**k / math.factorial(k)


def poisson_matrix(exp_home: float, exp_away: float, max_g: int = 5) -> list[list[float]]:
    ph = [poisson_pmf(i, exp_home) for i in range(max_g + 1)]
    pa = [poisson_pmf(i, exp_away) for i in range(max_g + 1)]
    return [[ph[h] * pa[a] for a in range(max_g + 1)] for h in range(max_g + 1)]


def poisson_most_likely(exp_home: float, exp_away: float) -> tuple[int, int]:
    mat = poisson_matrix(exp_home, exp_away)
    best_h, best_a, best_p = 0, 0, 0.0
    for h, row in enumerate(mat):
        for a, p in enumerate(row):
            if p > best_p:
                best_p, best_h, best_a = p, h, a
    return best_h, best_a


def poisson_over_total(exp_home: float, exp_away: float, line: float) -> float:
    mat = poisson_matrix(exp_home, exp_away)
    p_over = 0.0
    for h, row in enumerate(mat):
        for a, p in enumerate(row):
            if h + a > line:
                p_over += p
    return p_over * 100


def poisson_1x2_probs(exp_home: float, exp_away: float) -> tuple[float, float, float]:
    mat = poisson_matrix(exp_home, exp_away)
    p_h = p_d = p_a = 0.0
    for h, row in enumerate(mat):
        for a, p in enumerate(row):
            if h > a:
                p_h += p
            elif h == a:
                p_d += p
            else:
                p_a += p
    total = p_h + p_d + p_a or 1.0
    return p_h / total * 100, p_d / total * 100, p_a / total * 100


def _ou_label(expected: float | None, line: float) -> str:
    if expected is None:
        return ""
    return "over" if expected > line else "under"


def _confidence_score(
    gap: float,
    min_n: int,
    has_stats: bool,
    *,
    p_max: float = 0.0,
) -> int:
    if min_n < MIN_MATCHES:
        return max(0, min_n * 10)
    score = 40.0
    score += min(30, gap * 120)
    score += min(20, (min_n - MIN_MATCHES) * 4)
    score += min(10, (p_max - 0.33) * 30) if p_max > 0.33 else 0
    if not has_stats:
        score *= 0.85
    return int(max(0, min(100, round(score))))


def _confidence_label(score: int) -> str:
    if score >= 70:
        return "wysoka"
    if score >= 45:
        return "srednia"
    return "niska"


def compute_league_home_advantage(df: pd.DataFrame) -> dict[str, float]:
    out: dict[str, float] = {}
    for liga in df[COL_LIGA].dropna().unique():
        sub = df[(df[COL_LIGA] == liga) & df["_hg"].notna() & df["_ag"].notna()]
        if sub.empty:
            out[str(liga)] = base.HOME_ADVANTAGE
            continue
        hw = (sub["_hg"] > sub["_ag"]).mean()
        out[str(liga)] = max(1.0, min(1.25, 0.8 + hw * 0.8))
    return out


def compute_h2h(
    df: pd.DataFrame,
    home: str,
    away: str,
    *,
    as_of: pd.Timestamp | None,
    n: int = 5,
) -> dict[str, Any]:
    mask = (
        ((df[COL_HOME] == home) & (df[COL_AWAY] == away))
        | ((df[COL_HOME] == away) & (df[COL_AWAY] == home))
    )
    sub = df.loc[mask]
    if as_of is not None:
        sub = sub[sub[COL_DATA] < as_of]
    sub = sub.sort_values(COL_DATA).tail(n)
    if sub.empty:
        return {"h2h_mecze": 0, "h2h_home_wins": 0, "h2h_avg_goli": None, "h2h_btts_pct": None}

    home_wins = 0
    goals: list[float] = []
    btts = 0
    for _, m in sub.iterrows():
        g = base._team_goals(m, home)
        if g:
            if g[0] > g[1]:
                home_wins += 1
            goals.append(g[0] + g[1])
            if g[0] > 0 and g[1] > 0:
                btts += 1
        bval = str(m.get(COL_BTTS, "")).strip().casefold()
        if bval in {"yes", "tak", "так"}:
            btts += 1

    return {
        "h2h_mecze": len(sub),
        "h2h_home_wins": home_wins,
        "h2h_avg_goli": _mean(goals),
        "h2h_btts_pct": (btts / len(sub) * 100) if sub.shape[0] else None,
    }


def compute_league_rank(
    df: pd.DataFrame,
    team: str,
    league: str,
    *,
    as_of: pd.Timestamp | None,
) -> dict[str, Any]:
    sub = df[df[COL_LIGA] == league]
    if as_of is not None:
        sub = sub[sub[COL_DATA] < as_of]
    sub = sub[sub["_hg"].notna() & sub["_ag"].notna()]
    if sub.empty:
        return {"punkty": 0, "ranking": None, "mecze_ranking": 0}

    pts: dict[str, int] = {}
    for _, m in sub.iterrows():
        hg, ag = int(m["_hg"]), int(m["_ag"])
        for t in (m[COL_HOME], m[COL_AWAY]):
            pts.setdefault(str(t), 0)
        if hg > ag:
            pts[str(m[COL_HOME])] += 3
        elif hg < ag:
            pts[str(m[COL_AWAY])] += 3
        else:
            pts[str(m[COL_HOME])] += 1
            pts[str(m[COL_AWAY])] += 1

    sorted_teams = sorted(pts.items(), key=lambda x: (-x[1], x[0]))
    rank = next((i + 1 for i, (t, _) in enumerate(sorted_teams) if t == team), None)
    return {"punkty": pts.get(team, 0), "ranking": rank, "mecze_ranking": len(pts)}


def days_since_last(index: base.TeamStatsIndex, team: str, as_of: pd.Timestamp) -> int | None:
    recs = index._by_team.get(str(team), [])
    recs = [r for r in recs if r["date"] < as_of]
    if not recs:
        return None
    delta = (as_of - recs[-1]["date"]).days
    return int(delta)


class LeagueCalibration:
    """Bias O/U per liga (warstwa 2) — proste dostrojenie na 2024."""

    def __init__(self) -> None:
        self.foul_lines: dict[str, float] = {}
        self.goal_bias: dict[str, float] = {}
        self.corner_bias: dict[str, float] = {}
        self.card_bias: dict[str, float] = {}

    @classmethod
    def fit(cls, df: pd.DataFrame) -> LeagueCalibration:
        cal = cls()
        train = df[df[COL_DATA].dt.year == 2024].copy()
        for liga in train[COL_LIGA].dropna().unique():
            sub = train[train[COL_LIGA] == liga]
            if "фоли" in sub.columns:
                fouls = pd.to_numeric(sub["фоли"], errors="coerce").dropna()
                if not fouls.empty:
                    cal.foul_lines[str(liga)] = float(fouls.median())
            if "кутові" in sub.columns:
                c = pd.to_numeric(sub["кутові"], errors="coerce").dropna()
                if not c.empty:
                    cal.corner_bias[str(liga)] = float(c.mean() - c.median())
            if "жовті_картки" in sub.columns:
                k = pd.to_numeric(sub["жовті_картки"], errors="coerce").dropna()
                if not k.empty:
                    cal.card_bias[str(liga)] = float(k.mean() - k.median())
            g = sub["_hg"] + sub["_ag"]
            g = g.dropna()
            if not g.empty:
                cal.goal_bias[str(liga)] = float(g.mean() - g.median())
        return cal

    def adjust_corners(self, league: str, exp: float | None) -> float | None:
        if exp is None:
            return None
        return exp + self.corner_bias.get(league, 0.0)

    def adjust_cards(self, league: str, exp: float | None) -> float | None:
        if exp is None:
            return None
        return exp + self.card_bias.get(league, 0.0)

    def foul_line(self, league: str) -> float:
        return self.foul_lines.get(league, 21.0)


def build_features(
    home: str,
    away: str,
    league: str,
    df_history: pd.DataFrame,
    *,
    as_of: pd.Timestamp | None,
    index: base.TeamStatsIndex,
    team_avg: pd.DataFrame,
    home_adv: dict[str, float],
) -> dict[str, Any]:
    feats: dict[str, Any] = {"ліга": league, "господар": home, "гість": away}
    if as_of is not None:
        feats["as_of"] = as_of.strftime("%d/%m/%Y")

    def season_row(team: str) -> dict[str, Any]:
        hit = team_avg[team_avg["druzyna"] == team] if not team_avg.empty else pd.DataFrame()
        return hit.iloc[0].to_dict() if not hit.empty else {}

    h_s, a_s = season_row(home), season_row(away)
    for n in FORM_WINDOWS:
        hf = base.compute_form(df_history, home, as_of=as_of, n=n, index=index)
        af = base.compute_form(df_history, away, as_of=as_of, n=n, index=index)
        feats[f"forma_{n}_h_gf"] = hf.get("srednia_goli_strzelonych")
        feats[f"forma_{n}_a_gf"] = af.get("srednia_goli_strzelonych")
        feats[f"forma_{n}_h_mecze"] = hf.get("mecze", 0)
        feats[f"forma_{n}_a_mecze"] = af.get("mecze", 0)

    feats.update({f"h2h_{k}": v for k, v in compute_h2h(df_history, home, away, as_of=as_of).items()})
    rh = compute_league_rank(df_history, home, league, as_of=as_of)
    ra = compute_league_rank(df_history, away, league, as_of=as_of)
    feats["ranking_h"] = rh.get("ranking")
    feats["ranking_a"] = ra.get("ranking")
    feats["punkty_h"] = rh.get("punkty")
    feats["punkty_a"] = ra.get("punkty")
    if as_of is not None:
        feats["days_since_h"] = days_since_last(index, home, as_of)
        feats["days_since_a"] = days_since_last(index, away, as_of)
    feats["home_advantage_ligi"] = home_adv.get(league, base.HOME_ADVANTAGE)
    feats["ma_statystyki"] = has_league_stats(league, df_history)
    return feats


def predict_match_max(
    home: str,
    away: str,
    league: str,
    df_history: pd.DataFrame,
    team_avg: pd.DataFrame | None = None,
    *,
    as_of: pd.Timestamp | None = None,
    index: base.TeamStatsIndex | None = None,
    calibration: LeagueCalibration | None = None,
    home_adv: dict[str, float] | None = None,
) -> dict[str, Any]:
    hist = df_history
    if team_avg is None:
        team_avg = base.compute_team_averages(
            hist if as_of is None else hist[hist[COL_DATA] < as_of]
        )
    idx = index or base.get_team_index(df_history)
    cal = calibration or LeagueCalibration()
    ha = home_adv or compute_league_home_advantage(hist)
    has_stats = has_league_stats(league, hist)

    def season_row(team: str) -> dict[str, Any]:
        hit = team_avg[team_avg["druzyna"] == team] if not team_avg.empty else pd.DataFrame()
        return hit.iloc[0].to_dict() if not hit.empty else {}

    h_s, a_s = season_row(home), season_row(away)
    h_form = base.compute_form(hist, home, as_of=as_of, n=5, index=idx)
    a_form = base.compute_form(hist, away, as_of=as_of, n=5, index=idx)
    h_home = base.compute_form(hist, home, as_of=as_of, n=5, venue="home", index=idx)
    a_away = base.compute_form(hist, away, as_of=as_of, n=5, venue="away", index=idx)

    def pick(form: dict, season: dict, key: str) -> float | None:
        return base._blend(form.get(key), season.get(key))

    home_gf = base._blend(
        h_home.get("srednia_goli_strzelonych") or h_form.get("srednia_goli_strzelonych"),
        h_s.get("srednia_goli_strzelonych"),
    )
    home_ga = base._blend(
        h_home.get("srednia_goli_straconych") or h_form.get("srednia_goli_straconych"),
        h_s.get("srednia_goli_straconych"),
    )
    away_gf = base._blend(
        a_away.get("srednia_goli_strzelonych") or a_form.get("srednia_goli_strzelonych"),
        a_s.get("srednia_goli_strzelonych"),
    )
    away_ga = base._blend(
        a_away.get("srednia_goli_straconych") or a_form.get("srednia_goli_straconych"),
        a_s.get("srednia_goli_straconych"),
    )

    home_n = int(h_form.get("mecze") or h_s.get("mecze") or 0)
    away_n = int(a_form.get("mecze") or a_s.get("mecze") or 0)
    home_n_goals = int(h_s.get("mecze_z_wynikiem") or 0)
    away_n_goals = int(a_s.get("mecze_z_wynikiem") or 0)

    skip = home_n < MIN_MATCHES or away_n < MIN_MATCHES
    adv = ha.get(league, base.HOME_ADVANTAGE)

    use_goals = (
        home_gf is not None
        and away_gf is not None
        and home_ga is not None
        and away_ga is not None
        and (home_n_goals >= 2 or (h_form.get("mecze") or 0) >= 2)
        and (away_n_goals >= 2 or (a_form.get("mecze") or 0) >= 2)
    )

    if use_goals:
        metoda = base.METODA_FORMA if (h_form.get("mecze") or 0) >= 3 and (a_form.get("mecze") or 0) >= 3 else base.METODA_GOALS
        exp_home = max(0.2, (home_gf + away_ga) / 2 * adv)
        exp_away = max(0.2, (away_gf + home_ga) / 2)
        exp_home += cal.goal_bias.get(league, 0.0) / 2
        exp_away += cal.goal_bias.get(league, 0.0) / 2
    else:
        metoda = base.METODA_STATS
        league_sub = hist[hist[COL_LIGA] == league]
        league_sot = league_sub["удари_в_площину"].mean() if not league_sub.empty else 8.0
        if pd.isna(league_sot) or league_sot <= 0:
            league_sot = 8.0

        def sot(form: dict, season: dict) -> float:
            v = pick(form, season, "srednia_strzaly_w_stwor")
            return float(v) if v and v > 0 else float(league_sot)

        exp_home = max(0.2, sot(h_form, h_s) * adv * base.SOT_TO_GOALS)
        exp_away = max(0.2, sot(a_form, a_s) * base.SOT_TO_GOALS)

    p_home, p_draw, p_away = poisson_1x2_probs(exp_home, exp_away)
    p_max = max(p_home, p_draw, p_away) / 100

    if skip:
        winner, status, powod = "", "pomin", "za_malo_danych"
    elif not has_stats and metoda == base.METODA_STATS and league in NORDIC_LEAGUES:
        status, powod = "ok", "nordic_bez_stat"
        if p_home >= p_away and p_home >= p_draw:
            winner = home
        elif p_away >= p_home and p_away >= p_draw:
            winner = away
        else:
            winner = "remis"
    else:
        status, powod = "ok", ""
        if p_home >= p_away and p_home >= p_draw:
            winner = home
        elif p_away >= p_home and p_away >= p_draw:
            winner = away
        else:
            winner = "remis"

    gap = (max(p_home, p_away, p_draw) - sorted([p_home, p_away, p_draw])[1]) / 100
    conf_score = _confidence_score(gap, min(home_n, away_n), has_stats, p_max=p_max)
    pewnosc = _confidence_label(conf_score)

    pois_h, pois_a = poisson_most_likely(exp_home, exp_away)
    if winner == home and pois_h <= pois_a:
        pois_h = pois_a + 1
    elif winner == away and pois_a <= pois_h:
        pois_a = pois_h + 1
    elif winner == "remis" and pois_h != pois_a:
        pois_h = pois_a = max(pois_h, pois_a)

    h_btts = pick(h_form, h_s, "btts_yes_pct") or 50.0
    a_btts = pick(a_form, a_s, "btts_yes_pct") or 50.0
    h2h = compute_h2h(hist, home, away, as_of=as_of)
    if h2h.get("h2h_btts_pct") is not None and h2h["h2h_mecze"] >= 2:
        btts_pct = (h_btts + a_btts + h2h["h2h_btts_pct"]) / 3
    else:
        btts_pct = (h_btts + a_btts) / 2
    btts_pred = "так" if btts_pct >= 50 else "ні"

    h_corners = pick(h_form, h_s, "srednia_rozne")
    a_corners = pick(a_form, a_s, "srednia_rozne")
    exp_corners = cal.adjust_corners(league, (h_corners + a_corners) if h_corners and a_corners else None)
    h_cards = pick(h_form, h_s, "srednia_zolte_kartki")
    a_cards = pick(a_form, a_s, "srednia_zolte_kartki")
    exp_cards = cal.adjust_cards(league, (h_cards + a_cards) if h_cards and a_cards else None)
    h_fouls = pick(h_form, h_s, "srednia_faule")
    a_fouls = pick(a_form, a_s, "srednia_faule")
    exp_fouls = (h_fouls + a_fouls) if h_fouls and a_fouls else None
    foul_line = cal.foul_line(league)

    def rnd(v: float | None) -> float | None:
        return round(float(v), 2) if v is not None else None

    out: dict[str, Any] = {
        "status": status,
        "powod": powod,
        "przewidywany_zwyciezca": winner,
        "prawdopodobienstwo_gospodarz": round(p_home, 1),
        "prawdopodobienstwo_remis": round(p_draw, 1),
        "prawdopodobienstwo_gosc": round(p_away, 1),
        "przewidywany_wynik": f"{pois_h}:{pois_a}",
        "przewidywany_wynik_poisson": f"{pois_h}:{pois_a}",
        "exp_gole_gospodarz": rnd(exp_home),
        "exp_gole_gosc": rnd(exp_away),
        "przewidywane_btts": btts_pred,
        "przewidywane_btts_pct": round(btts_pct, 1),
        "exp_rozne": rnd(exp_corners) if has_stats else None,
        "exp_zolte": rnd(exp_cards) if has_stats else None,
        "exp_faule": rnd(exp_fouls) if has_stats else None,
        "linia_faule": round(foul_line, 1) if has_stats else None,
        "forma_gospodarz_mecze": h_form.get("mecze", 0),
        "forma_gosc_mecze": a_form.get("mecze", 0),
        "metoda": metoda,
        "pewnosc": pewnosc,
        "pewnosc_score": conf_score,
        "min_meczy_historii": min(home_n, away_n),
        "ma_statystyki": "так" if has_stats else "ні",
    }

    for line in GOAL_LINES:
        p_over = poisson_over_total(exp_home, exp_away, line)
        out[f"ou_gole_{line}_pct"] = round(p_over, 1)
        out[f"predykcja_gole_{line}"] = "over" if p_over >= 50 else "under"

    if has_stats:
        for line in CORNER_LINES:
            out[f"predykcja_rozne_{line}"] = _ou_label(exp_corners, line)
        for line in CARD_LINES:
            out[f"predykcja_zolte_{line}"] = _ou_label(exp_cards, line)
        out["predykcja_rozne"] = out.get("predykcja_rozne_9.5", "")
        out["predykcja_zolte"] = out.get("predykcja_zolte_3.5", "")
        out["linia_rozne"] = 9.5
        out["linia_zolte"] = 3.5
        out["predykcja_faule"] = _ou_label(exp_fouls, foul_line)
        out["rozne_gospodarz_exp"] = rnd(h_corners)
        out["rozne_gosc_exp"] = rnd(a_corners)
        out["pred_rozne_gospodarz_4_5"] = _ou_label(h_corners, 4.5)
        out["pred_rozne_gosc_4_5"] = _ou_label(a_corners, 4.5)
    else:
        for line in CORNER_LINES:
            out[f"predykcja_rozne_{line}"] = ""
        for line in CARD_LINES:
            out[f"predykcja_zolte_{line}"] = ""
        out["predykcja_rozne"] = ""
        out["predykcja_zolte"] = ""
        out["predykcja_faule"] = ""

    # Handicap
    out["handicap_0"] = home if exp_home >= exp_away else ("remis" if abs(exp_home - exp_away) < 0.15 else away)
    out["handicap_h_minus_0_5"] = home if (p_home / 100 + 0.5 * p_draw / 100) > 0.5 else away

    # Value flag
    league_sub = hist[hist[COL_LIGA] == league]
    base_hw = (league_sub["_hg"] > league_sub["_ag"]).mean() if not league_sub.empty else 0.42
    edge = abs(p_home / 100 - base_hw)
    out["value_flag"] = "так" if conf_score >= 70 and edge >= 0.10 else "ні"

    return out


def build_predictions_max(
    df_target: pd.DataFrame,
    df_history: pd.DataFrame,
    *,
    use_as_of: bool = True,
    calibration: LeagueCalibration | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    cal = calibration or LeagueCalibration.fit(df_history)
    home_adv = compute_league_home_advantage(df_history)
    if use_as_of and not df_target.empty:
        cut = pd.Timestamp(df_target[COL_DATA].min())
        season_avg = base.compute_team_averages(df_history, as_of=cut)
    else:
        season_avg = base.compute_team_averages(df_history)
    index = base.get_team_index(df_history)

    rows: list[dict[str, Any]] = []
    feat_rows: list[dict[str, Any]] = []

    for _, m in df_target.iterrows():
        as_of = pd.Timestamp(m[COL_DATA]) if use_as_of else None
        pred = predict_match_max(
            m[COL_HOME],
            m[COL_AWAY],
            m[COL_LIGA],
            df_history,
            team_avg=season_avg,
            as_of=as_of,
            index=index,
            calibration=cal,
            home_adv=home_adv,
        )
        actual = str(m[COL_RESULT]).strip() if COL_RESULT in m.index and pd.notna(m[COL_RESULT]) else ""
        rows.append(
            {
                COL_LIGA: m[COL_LIGA],
                COL_DATA: m[COL_DATA],
                COL_HOME: m[COL_HOME],
                COL_AWAY: m[COL_AWAY],
                COL_RESULT: actual,
                **pred,
            }
        )
        if len(feat_rows) < 100:
            feats = build_features(
                m[COL_HOME],
                m[COL_AWAY],
                m[COL_LIGA],
                df_history,
                as_of=as_of,
                index=index,
                team_avg=season_avg,
                home_adv=home_adv,
            )
            feat_rows.append(feats)

    out = pd.DataFrame(rows)
    features = pd.DataFrame(feat_rows)
    if not out.empty:
        out[COL_DATA] = pd.to_datetime(out[COL_DATA]).dt.strftime("%d/%m/%Y")
    return out, features


def _actual_total_goals(row: pd.Series) -> float | None:
    score = base.parse_score(row.get(COL_RESULT))
    return float(score[0] + score[1]) if score else None


def run_backtest_max(
    df_history: pd.DataFrame,
    *,
    year: int = 2025,
    calibration: LeagueCalibration | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    target = df_history[df_history[COL_DATA].dt.year == year].copy()
    if target.empty:
        return pd.DataFrame(), pd.DataFrame()

    cal = calibration or LeagueCalibration.fit(df_history[df_history[COL_DATA].dt.year <= 2024])
    preds, _ = build_predictions_max(target, df_history, use_as_of=True, calibration=cal)
    detail = preds.copy()
    detail["actual_zwyciezca"] = [
        base._actual_winner(r) for _, r in target.reset_index(drop=True).iterrows()
    ]
    detail["actual_btts"] = [
        base._actual_btts(r) for _, r in target.reset_index(drop=True).iterrows()
    ]
    detail["actual_rozne"] = pd.to_numeric(target["кутові"].values, errors="coerce") if "кутові" in target.columns else pd.NA
    detail["actual_zolte"] = (
        pd.to_numeric(target["жовті_картки"].values, errors="coerce")
        if "жовті_картки" in target.columns
        else pd.NA
    )
    detail["actual_gole"] = [_actual_total_goals(r) for _, r in target.reset_index(drop=True).iterrows()]

    scored = detail[detail["status"] == "ok"].copy()
    scored = scored[scored["actual_zwyciezca"].astype(str).str.len() > 0]

    def hit_rate(mask: pd.Series, correct: pd.Series) -> float | None:
        sub = correct[mask]
        return round(float(sub.mean() * 100), 1) if len(sub) else None

    if scored.empty:
        return detail, pd.DataFrame([{"metryka": "brak_danych", "wartosc": 0}])

    ok_1x2 = scored["przewidywany_zwyciezca"] == scored["actual_zwyciezca"]
    ok_btts = scored["przewidywane_btts"] == scored["actual_btts"]

    has_corners = scored["actual_rozne"].notna() & (scored["predykcja_rozne"].astype(str) != "")
    actual_corners_ou = scored["actual_rozne"].map(
        lambda x: "over" if pd.notna(x) and float(x) > 9.5 else ("under" if pd.notna(x) else "")
    )
    ok_corners = scored["predykcja_rozne"] == actual_corners_ou

    has_cards = scored["actual_zolte"].notna() & (scored["predykcja_zolte"].astype(str) != "")
    actual_cards_ou = scored["actual_zolte"].map(
        lambda x: "over" if pd.notna(x) and float(x) > 3.5 else ("under" if pd.notna(x) else "")
    )
    ok_cards = scored["predykcja_zolte"] == actual_cards_ou

    has_goals = scored["actual_gole"].notna() & scored["predykcja_gole_2.5"].astype(str).ne("")
    actual_goals_ou = scored["actual_gole"].map(
        lambda x: "over" if pd.notna(x) and float(x) > 2.5 else ("under" if pd.notna(x) else "")
    )
    ok_goals_25 = scored["predykcja_gole_2.5"] == actual_goals_ou

    summary_rows = [
        {"metryka": "rok", "wartosc": year},
        {"metryka": "model", "wartosc": "predykcje_max"},
        {"metryka": "mecze_razem", "wartosc": len(detail)},
        {"metryka": "mecze_ocenione", "wartosc": len(scored)},
        {"metryka": "hit_rate_1x2_pct", "wartosc": round(float(ok_1x2.mean() * 100), 1)},
        {"metryka": "hit_rate_btts_pct", "wartosc": round(float(ok_btts.mean() * 100), 1)},
        {"metryka": "hit_rate_gole_2_5_pct", "wartosc": hit_rate(has_goals, ok_goals_25) if has_goals.any() else None},
        {"metryka": "hit_rate_rozne_9_5_pct", "wartosc": hit_rate(has_corners, ok_corners) if has_corners.any() else None},
        {"metryka": "hit_rate_zolte_3_5_pct", "wartosc": hit_rate(has_cards, ok_cards) if has_cards.any() else None},
    ]

    detail["trafiony_1x2"] = (detail["przewidywany_zwyciezca"] == detail["actual_zwyciezca"]).map(
        {True: "так", False: "ні"}
    )
    return detail, pd.DataFrame(summary_rows)


def run_benchmark(df_history: pd.DataFrame, *, year: int = 2025) -> pd.DataFrame:
    """Porównanie predykcje_max vs predykcje.py na roku testowym."""
    _, base_summary = base.run_backtest(df_history, year=year)
    _, max_summary = run_backtest_max(df_history, year=year)

    metrics = [
        "hit_rate_1x2_pct",
        "hit_rate_btts_pct",
        "hit_rate_rozne_9_5_pct",
        "hit_rate_zolte_3_5_pct",
    ]

    rows: list[dict[str, Any]] = []
    for m in metrics:
        b_val = base_summary.loc[base_summary["metryka"] == m, "wartosc"]
        m_val = max_summary.loc[max_summary["metryka"] == m, "wartosc"]
        bv = float(b_val.iloc[0]) if not b_val.empty and pd.notna(b_val.iloc[0]) else None
        mv = float(m_val.iloc[0]) if not m_val.empty and pd.notna(m_val.iloc[0]) else None
        delta = round(mv - bv, 1) if bv is not None and mv is not None else None
        rows.append(
            {
                "metryka": m,
                "baseline_predykcje": bv,
                "predykcje_max": mv,
                "delta_pp": delta,
                "lepszy": "max" if delta and delta > 0 else ("baseline" if delta and delta < 0 else "remis"),
            }
        )

    # dodatkowa metryka tylko w max
    g_val = max_summary.loc[max_summary["metryka"] == "hit_rate_gole_2_5_pct", "wartosc"]
    if not g_val.empty:
        rows.append(
            {
                "metryka": "hit_rate_gole_2_5_pct",
                "baseline_predykcje": None,
                "predykcje_max": float(g_val.iloc[0]) if pd.notna(g_val.iloc[0]) else None,
                "delta_pp": None,
                "lepszy": "max_only",
            }
        )
    return pd.DataFrame(rows)


def select_top_typy(predictions: pd.DataFrame, *, top_n: int = TOP_N_TYPOW) -> pd.DataFrame:
    sub = predictions[predictions["status"] == "ok"].copy()
    sub = sub[sub["pewnosc"].isin(["wysoka", "srednia"])]
    if sub.empty:
        return sub
    sub = sub.sort_values("pewnosc_score", ascending=False).head(top_n)
    return sub


def _load_audyt_df() -> pd.DataFrame:
    if not AUDYT_MD.exists():
        return pd.DataFrame([{"інформація": "Немає AUDYT_DANYCH.md — запустіть scripts/audyt_danych.py"}])
    text = AUDYT_MD.read_text(encoding="utf-8")
    lines = [ln for ln in text.splitlines() if ln.startswith("|") and "---" not in ln]
    if len(lines) < 2:
        return pd.DataFrame([{"інформація": text[:500]}])
    header = [c.strip() for c in lines[0].strip("|").split("|")]
    rows = []
    for ln in lines[1:]:
        rows.append([c.strip() for c in ln.strip("|").split("|")])
    return pd.DataFrame(rows, columns=header[: len(rows[0])])


def _write_table_sheet(
    wb: Workbook,
    title: str,
    df: pd.DataFrame,
    *,
    table_names: set[str],
    translate: bool = True,
) -> None:
    """Arkusz z danymi owiniętymi w tabelę Excel; domyślnie kolumny/wartości UA."""
    ws = wb.create_sheet(title=title[:31])
    export_df = df.drop(columns=[c for c in ("_hg", "_ag") if c in df.columns], errors="ignore")
    export_df = drop_empty_columns(export_df)
    if translate:
        export_df = translate_df_ua(export_df)
    if export_df.empty:
        ws.append(["немає_даних"])
        return
    for r in dataframe_to_rows(export_df, index=False, header=True):
        ws.append(list(r))
    if ws.max_row >= 2 and ws.max_column >= 1:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        tab = Table(displayName=_table_display_name(title, table_names), ref=ref)
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tab)
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 16


def _safe_sheet_name(name: str, used: set[str]) -> str:
    clean = re.sub(r"[\[\]:*?/\\]", "", name)[:25]
    base_name = clean or "Liga"
    candidate = base_name
    i = 1
    while candidate in used:
        suffix = f"_{i}"
        candidate = base_name[: 31 - len(suffix)] + suffix
        i += 1
    used.add(candidate)
    return candidate


def export_excel_max(
    df_2026: pd.DataFrame,
    predictions: pd.DataFrame,
    features: pd.DataFrame,
    team_avg: pd.DataFrame,
    league_avg: pd.DataFrame,
    path: Path | None = None,
    *,
    monthly: pd.DataFrame | None = None,
    backtest_detail: pd.DataFrame | None = None,
    backtest_summary: pd.DataFrame | None = None,
    benchmark: pd.DataFrame | None = None,
    top_typy: pd.DataFrame | None = None,
) -> Path:
    out_path = path or OUT_XLSX
    wb = Workbook()
    wb.remove(wb.active)
    table_names: set[str] = set()
    used_sheets: set[str] = set()

    leagues = sorted(df_2026[COL_LIGA].dropna().unique())

    # --- Osobny arkusz per liga (prognozy 2026) ---
    pred_ua = predictions.copy()
    if COL_LIGA in pred_ua.columns:
        for league in leagues:
            sub = pred_ua[pred_ua[COL_LIGA] == league].copy()
            if sub.empty:
                continue
            sub = sub.drop(columns=[COL_LIGA], errors="ignore")
            sheet = _safe_sheet_name(str(league), used_sheets)
            _write_table_sheet(wb, sheet, sub, table_names=table_names)
    elif not pred_ua.empty:
        _write_table_sheet(wb, "Прогнози", pred_ua, table_names=table_names)

    if top_typy is not None and not top_typy.empty:
        _write_table_sheet(wb, UA_SHEET_SUMMARY["top_typy"], top_typy, table_names=table_names)

    if features is not None and not features.empty:
        for league in leagues:
            if COL_LIGA not in features.columns and "ліга" not in features.columns:
                break
            lcol = COL_LIGA if COL_LIGA in features.columns else "ліга"
            fsub = features[features[lcol] == league].copy()
            if fsub.empty:
                continue
            fsub = fsub.drop(columns=[lcol], errors="ignore")
            sheet = _safe_sheet_name(f"Ознаки_{league}", used_sheets)
            _write_table_sheet(wb, sheet, fsub, table_names=table_names)

    _write_table_sheet(wb, UA_SHEET_SUMMARY["team_avg"], team_avg, table_names=table_names)
    _write_table_sheet(wb, UA_SHEET_SUMMARY["league_avg"], league_avg, table_names=table_names)

    if monthly is None:
        monthly = compute_monthly_team_summary(df_2026, year=2026)
    if monthly is not None and not monthly.empty:
        _write_table_sheet(wb, UA_SHEET_SUMMARY["monthly"], monthly, table_names=table_names)

    if backtest_summary is not None and not backtest_summary.empty:
        _write_table_sheet(wb, UA_SHEET_SUMMARY["backtest"], backtest_summary, table_names=table_names)

    if backtest_detail is not None and not backtest_detail.empty:
        if COL_LIGA in backtest_detail.columns:
            for league in sorted(backtest_detail[COL_LIGA].dropna().unique()):
                bsub = backtest_detail[backtest_detail[COL_LIGA] == league].copy()
                bsub = bsub.drop(columns=[COL_LIGA], errors="ignore")
                sheet = _safe_sheet_name(f"Бектест_{league}", used_sheets)
                _write_table_sheet(wb, sheet, bsub, table_names=table_names)
        else:
            _write_table_sheet(
                wb, UA_SHEET_SUMMARY["backtest_detail"], backtest_detail, table_names=table_names
            )

    if benchmark is not None and not benchmark.empty:
        _write_table_sheet(wb, UA_SHEET_SUMMARY["benchmark"], benchmark, table_names=table_names)

    _write_table_sheet(wb, UA_SHEET_SUMMARY["audyt"], _load_audyt_df(), table_names=table_names, translate=False)

    # --- Wykresy (ukraińskie etykiety) ---
    ws_charts = wb.create_sheet(title=UA_SHEET_SUMMARY["charts"][:31])
    ws_charts["A1"] = "Кругові діаграми — частка прогнозованих перемог по командах"
    chart_row = 3
    for league in leagues:
        sub = predictions[(predictions[COL_LIGA] == league) & (predictions.get("status", "ok") != "pomin")]
        if sub.empty:
            continue
        wins: dict[str, int] = {}
        for _, r in sub.iterrows():
            w = r.get("przewidywany_zwyciezca", "")
            if not w or w == "remis":
                continue
            wins[w] = wins.get(w, 0) + 1
        if len(wins) < 2:
            continue
        total = sum(wins.values())
        pie_df = pd.DataFrame(
            [{"команда": k, "перемоги_відсоток": round(v / total * 100, 1)} for k, v in sorted(wins.items())]
        )
        start_row = chart_row
        ws_charts.cell(row=start_row, column=1, value=f"Ліга: {league}")
        ws_charts.cell(row=start_row + 1, column=1, value="команда")
        ws_charts.cell(row=start_row + 1, column=2, value="перемоги_відсоток")
        for i, row in pie_df.iterrows():
            ws_charts.cell(row=start_row + 2 + i, column=1, value=row["команда"])
            ws_charts.cell(row=start_row + 2 + i, column=2, value=row["перемоги_відсоток"])
        data_end = start_row + 1 + len(pie_df)
        if data_end > start_row + 1:
            ref = f"A{start_row + 1}:B{data_end}"
            tab = Table(displayName=_table_display_name(f"Chart_{league}", table_names), ref=ref)
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            ws_charts.add_table(tab)
        chart = PieChart()
        chart.title = f"{league} — прогноз перемог (%)"
        labels = Reference(ws_charts, min_col=1, min_row=start_row + 2, max_row=data_end)
        data = Reference(ws_charts, min_col=2, min_row=start_row + 1, max_row=data_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.width = 14
        chart.height = 10
        ws_charts.add_chart(chart, f"D{start_row}")
        chart_row = data_end + 18

    wb.save(out_path)
    logger.info("Zapisano Excel MAX: %s", out_path)
    return out_path


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Predykcje MAX")
    parser.add_argument("--ml", action="store_true", help="Warstwa 3 ML (opcjonalna)")
    args = parser.parse_args(argv)

    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    if args.ml:
        _safe_print("Warstwa 3 ML: pominieta (brak poprawy nad L1/L2 w tym buildzie)")

    src = resolve_aleks_path()
    if src is None:
        raise SystemExit(f"Brak pliku: {ALEKS_XLSX}")

    df_2026, df_history = base.load_2026_data(src)
    _safe_print(
        f"Mecze 2026: {len(df_2026)} | historii: {len(df_history)}"
    )
    if df_2026.empty:
        raise SystemExit("Brak danych z 2026.")

    cal = LeagueCalibration.fit(df_history)
    team_avg = base.compute_team_averages(df_history)
    league_avg = base.compute_league_averages(df_2026)

    _safe_print("Buduje predykcje MAX 2026...")
    predictions, features = build_predictions_max(df_2026, df_history, calibration=cal)
    top_typy = select_top_typy(predictions)

    _safe_print("Backtest 2025 + benchmark vs baseline...")
    bt_detail, bt_summary = run_backtest_max(df_history, year=2025, calibration=cal)
    benchmark = run_benchmark(df_history, year=2025)

    if not bt_summary.empty:
        for _, r in bt_summary.iterrows():
            _safe_print(f"  {r['metryka']}: {r['wartosc']}")
    if not benchmark.empty:
        _safe_print("Benchmark (max vs baseline):")
        for _, r in benchmark.iterrows():
            _safe_print(f"  {r['metryka']}: delta {r.get('delta_pp')} pp")

    path = export_excel_max(
        df_2026,
        predictions,
        features,
        team_avg,
        league_avg,
        backtest_detail=bt_detail,
        backtest_summary=bt_summary,
        benchmark=benchmark,
        top_typy=top_typy,
    )

    _safe_print(f"Predykcje MAX: {len(predictions)} meczy | Top_typy: {len(top_typy)}")
    _safe_print(f"Plik: {path}")


if __name__ == "__main__":
    main()
